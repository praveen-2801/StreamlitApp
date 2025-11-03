from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO
from copy import copy
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import os, shutil
from io import BytesIO
from copy import copy as _copy
from openpyxl import load_workbook
import math


# ----------------- PAGE CONFIGURATION ----------------- #
st.set_page_config(page_title="Pricing Tool", layout="wide")

# ----------------- SIDEBAR NAVIGATION ----------------- #
st.sidebar.title("Navigation")
selected_option = st.sidebar.radio("Choose an option", ["Home", "Pricing Analysis Version - 2"])

# ----------------- FILE UPLOAD FUNCTIONS ----------------- #
def TCP_info_upload_files():
    TCP_uploaded = st.file_uploader("Upload TCP information file", type=["xlsx"], accept_multiple_files=False)
    if TCP_uploaded:
        st.success("TCP information file uploaded successfully.")
    return TCP_uploaded

def upload_files():
    uploaded = st.file_uploader("Upload Supplier Excel files", type=["xlsx"], accept_multiple_files=True)
    if uploaded:
        st.success(f"{len(uploaded)} file(s) uploaded successfully.")
    return uploaded

# ----------------- EXTRACT DATA FROM SUPPLIER FILE ----------------- #
def extract_data_by_reference(file):
    in_memory_file = BytesIO(file.read())
    wb = load_workbook(in_memory_file, data_only=False)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))

    merged_ranges = ws.merged_cells.ranges
    title_row = [{
        "value": cell.value,
        "fill": PatternFill(fill_type=None),
        "font": copy(cell.font),
        "alignment": copy(cell.alignment),
        "border": copy(cell.border)
    } for cell in rows[0]]

    headers = [cell.value for cell in rows[1]]
    ref_col_idx = next((idx for idx, val in enumerate(headers)
                        if val and ("tcp" in str(val).lower() or "reference number" in str(val).lower())), None)
    if ref_col_idx is None:
        raise ValueError("Neither 'TCP Number' nor 'Reference Number' column found")

    header_cells = [{
        "value": cell.value,
        "fill": PatternFill(fill_type=None),
        "font": copy(cell.font),
        "alignment": copy(cell.alignment),
        "border": copy(cell.border)
    } for cell in rows[1]]

    formatted_data = {}
    for row in rows[2:]:
        ref_val = row[ref_col_idx].value
        if not ref_val:
            continue
        cell_data_row = [{
            "value": cell.value,
            "fill": PatternFill(fill_type=None),
            "font": copy(cell.font),
            "alignment": copy(cell.alignment),
            "border": copy(cell.border)
        } for cell in row]
        formatted_data[ref_val] = cell_data_row

    return title_row, header_cells, ref_col_idx, formatted_data, merged_ranges

# ----------------- EXTRACT TCP DATA BY REFERENCE ----------------- #
def extract_tcp_data_by_reference(file):
    in_memory_file = BytesIO(file.read())
    wb = load_workbook(in_memory_file, data_only=False)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))

    # Clean headers (strip spaces)
    headers = [str(cell.value).strip() if cell.value else "" for cell in rows[0]]

    # Detect reference column
    ref_col_idx = next(
        (idx for idx, val in enumerate(headers)
         if val and ("tcp" in val.lower() or "reference" in val.lower())),
        None
    )
    if ref_col_idx is None:
        raise ValueError("Neither 'TCP Number' nor 'Reference Number' column found in TCP file")

    # Collect formatted rows into dict {ref: row_data}
    formatted_data = {}
    for row in rows[1:]:
        ref_val = str(row[ref_col_idx].value).strip() if row[ref_col_idx].value else None
        if not ref_val:
            continue

        row_data = {}
        for idx, cell in enumerate(row):
            col_header = headers[idx]
            row_data[col_header] = cell.value

        formatted_data[ref_val] = row_data

    return headers, ref_col_idx, formatted_data

# ----------------- MERGE SUPPLIER FILES ----------------- #
def merge_file_by_reference(uploaded_files, tcp_file):
    all_references = set()
    file_data = []
    title_rows = []
    file_names = []

    for file in uploaded_files:
        title_row, headers, ref_idx, data, merged_ranges = extract_data_by_reference(file)
        if not data:
            st.warning(f"No data found in file: {file.name}")
            continue
        title_rows.append((title_row, merged_ranges))
        file_data.append((headers, ref_idx, data))
        file_names.append(file.name)
        all_references.update([str(k).strip() for k in data.keys() if k])


   
    # Extract TCP data if provided
    tcp_data = None
    tcp_headers = None
    tcp_ref_idx = None
    if tcp_file:
        tcp_headers, tcp_ref_idx, tcp_data = extract_tcp_data_by_reference(tcp_file)
        # normalize TCP refs
        tcp_data = {str(k).strip(): v for k, v in tcp_data.items()}
        all_references.update(list(tcp_data.keys()))


    if not all_references:
        st.error("No references found in the uploaded files.")
        return [], [], [], [], None, None, None

    # Convert all references to strings for consistent sorting
    all_references = [str(ref) for ref in all_references]
   
    # Always keep references as strings
    sorted_refs = sorted(all_references, key=lambda x: str(x).strip())


    return sorted_refs, file_data, title_rows, file_names, tcp_headers, tcp_ref_idx, tcp_data

def get_tcp_inputs(context="default"):
    st.subheader("TCP Input Parameters")
   
    cols = st.columns(2)
    with cols[0]:
        # general_duty_pct = st.number_input("General duty percentage (%)", min_value=0.0, value=0.0, format="%.2f", key=f"general_duty_pct_{context}")
        # reciprocal_pct = st.number_input("Reciprocal percentage (%)", min_value=0.0, value=0.0, format="%.2f", key=f"reciprocal_pct_{context}")
        Warehouse_inputs = st.radio(
            "Choose the ware house : ",
            ["ELM","TAL"],
        )
       
    with cols[1]:
        storage_months = st.number_input("Storage months", min_value=0, value=1,key=f"storage_months_{context}")
        markup = st.number_input("Markup percentage (%)", min_value=0.0, value=0.0, format="%.2f",key=f"markup_{context}")
        ftl_outfreight = st.number_input("FTL Outfreight value ($)", min_value=0.0, value=0.0, format="%.2f",key=f"ftl_outfreight_{context}")
        ltl_outfreight = st.number_input("LTL Outfreight value ($)", min_value=0.0, value=0.0, format="%.2f",key=f"ltl_outfreight_{context}")
   
    return {
        'storage_months': storage_months,
        'markup': markup,
        'FTL_Outfreight': ftl_outfreight,
        'LTL_Outfreight': ltl_outfreight,
        "Warehouse_inputs" : Warehouse_inputs,
    }

def calculate_tihi(carton_width, carton_length, carton_height, carton_weight,
                   pallet_width, pallet_length, max_pallet_height, max_pallet_weight):
    try:
        orientations = [
            (carton_width, carton_length),
            (carton_length, carton_width),
        ]

        best_total_cartons = 0

        def cases_per_layer(p_width, p_length, c_width, c_length):
            return (p_width // c_width) * (p_length // c_length)

        for c_width, c_length in orientations:
            ti = cases_per_layer(pallet_width, pallet_length, c_width, c_length)
            max_layers_by_height = max_pallet_height // carton_height
            max_cartons_by_weight = max_pallet_weight // carton_weight if carton_weight > 0 else 0
            total_cartons = min(ti * max_layers_by_height, max_cartons_by_weight)
            hi = total_cartons // ti if ti > 0 else 0

            if total_cartons > best_total_cartons:
                best_total_cartons = total_cartons

        return best_total_cartons
    except Exception as e:
        print(e)
        return 0

# ------------------ COMPARISON FUNCTION ------------------ #
def comparison_function(all_refs, file_data, file_names):
    results = {}
    default_supplier_name = file_names[0].replace(".xlsx", "")

    for ref in all_refs:
        fob_values = []
        best_supplier = default_supplier_name

        for idx, (headers, ref_idx, data) in enumerate(file_data):
            fob_col_idx = None
            for i, header in enumerate(headers):
                if i == ref_idx:
                    continue
                header_val = str(header["value"]).strip().lower() if header["value"] else ""
                if "fob price" in header_val and "case" in header_val:
                    fob_col_idx = i
                    break
            # Normalize keys in supplier data before lookup
            normalized_keys = {str(k).strip(): k for k in data.keys()}
            ref_str = str(ref).strip()

            if fob_col_idx is not None and ref_str in normalized_keys:
                key = normalized_keys[ref_str]
                try:
                    value = data[key][fob_col_idx]["value"]
                    if value is not None:
                        numeric_value = float(value)
                        fob_values.append(numeric_value)
                        if numeric_value < float('inf'):
                            best_supplier = file_names[idx].replace(".xlsx", "")
                except (ValueError, TypeError):
                    continue
        if fob_values:
            fob_values = [v for v in fob_values if v > 0]
            if fob_values:
                min_fob = round(min(fob_values), 2)
            else:
                min_fob = "N/A"
        else:
            min_fob = "N/A"

        results[ref] = {
            "Min FOB/Case": "$"+str(min_fob),
            "Min Supplier": best_supplier
        }

    return results


def calculation_for_min_fob_supplier(uploaded_files, tcp_file, tcp_inputs, min_fob_results):
    results = {}

    # Load TCP Information file
    tcp_df = pd.read_excel(tcp_file)

    # Pallet constants
    pallet_length = tcp_inputs.get('pallet_length', 40)
    pallet_width = tcp_inputs.get('pallet_width', 48)
    max_pallet_height = tcp_inputs.get('max_pallet_height', 85)
    max_pallet_weight = tcp_inputs.get('max_pallet_weight', 1200)
    inbound_to_east_ocean_freight_amount = tcp_inputs.get('ocean_freight', 5500)
    certification_cost = tcp_inputs.get('certification_cost', 1000)
    drayage_cost = tcp_inputs.get('drayage_cost', 1300)
    pallet_handling_cost = tcp_inputs.get('pallet_handling_cost', 50)
    storage_cost_per_pallet = tcp_inputs.get('storage_cost_per_pallet',28)
   
    testing_price = 0
    General_duty_price = 0
    # Reciprocal_Per_Case = 0
    Section_232_duty = tcp_inputs.get('section_232_duty', 25)
    chargeback_fee = tcp_inputs.get('chargeback_fee', 0)

    # Extract TCP inputs
    # Reciprocal_pct = tcp_inputs['reciprocal_pct']
    storage_months = tcp_inputs['storage_months']
    markup = tcp_inputs['markup']
    FTL_Outfreight = tcp_inputs['FTL_Outfreight']
    LTL_Outfreight = tcp_inputs['LTL_Outfreight']
    Warehouse_inputs = tcp_inputs["Warehouse_inputs"]
    selling_price = tcp_inputs.get('selling_price', 0)

    # Find reference column in TCP data
    tcp_ref_col = None
    for col in tcp_df.columns:
        if "tcp" in str(col).lower() or "reference" in str(col).lower() or "sr" in str(col).lower():
            tcp_ref_col = col
            break
   
    if tcp_ref_col is None:
        tcp_ref_col = tcp_df.columns[0]  # Use first column if no match

    for ref, min_supplier in min_fob_results.items():
        if min_supplier == "N/A":
            results[ref] = {"Error": "No valid supplier found"}
            continue
       
        # Find TCP row for this reference
        tcp_row_data = tcp_df[tcp_df[tcp_ref_col].astype(str).str.strip() == str(ref).strip()]
        tcp_row = tcp_row_data.to_dict("records")[0] if not tcp_row_data.empty else {}


        # Extract TCP values
        TI = float(tcp_row.get("TI", 0))
        HI = float(tcp_row.get("HI", 0))
        annual_vol = tcp_row.get("Annual Estimated Volume (Cases)",0)
        cases_fit = float(tcp_row.get("No Of Cases That Fit Into A Container", 0))
        General_duty_pct = float(tcp_row.get("General Duty", 0)) * 100
        Reciprocal_pct = float(tcp_row.get("Reciprocal", 0)) * 100
        Section_301_china = float(tcp_row.get("Section 301 China", 0))* 100
        Section_232_duty = float(tcp_row.get("Section 232 duty", 0)) * 100
        Tariff = float(tcp_row.get("Tariff", 0)) * 100
        Total_duty = float(tcp_row.get("Total Duty", 0)) * 100

        supplier_file_path = None
        for file in uploaded_files:
            if min_supplier in file.name:
                supplier_file_path = file
                break

        if supplier_file_path is None:
            raise FileNotFoundError(f"File not found for supplier '{min_supplier}'.")

        # Try reading first few rows to detect header dynamically
        preview = pd.read_excel(supplier_file_path, nrows=5, header=None)

        header_row_index = None
        for i in range(len(preview)):
            row_values = preview.iloc[i].astype(str).str.lower().tolist()
            if any("tcp number" in v for v in row_values) or any("reference number" in v for v in row_values):
                header_row_index = i
                break

        # Default fallback to first row
        if header_row_index is None:
            header_row_index = 0

        # Now read the file with detected header row
        df_supplier = pd.read_excel(supplier_file_path, header=header_row_index)
        df_supplier.columns = df_supplier.columns.map(lambda x: str(x).strip())

        # Select reference column
        if 'TCP Number' in df_supplier.columns:
            ref_col = 'TCP Number'
        elif 'Reference Number' in df_supplier.columns:
            ref_col = 'Reference Number'
        else:
            raise ValueError(
                f"No valid reference column ('TCP Number' or 'Reference Number') found. Columns: {df_supplier.columns.tolist()}"
            )

        # Normalize references before matching
        df_supplier[ref_col] = df_supplier[ref_col].astype(str).str.strip()
        ref_str = str(ref).strip()

        # Normalize supplier reference column and incoming ref
        df_supplier[ref_col] = df_supplier[ref_col].astype(str).str.strip()
        ref_str = str(ref).strip()

        if ref_str not in df_supplier[ref_col].values:
            raise ValueError(
                f"No data found for reference '{ref_str}' in {min_supplier}. "
                f"Available refs in {min_supplier}: {df_supplier[ref_col].unique()[:10]}"
            )

        df_filtered = df_supplier[df_supplier[ref_col] == ref_str]

       
        # Find carton dimensions
        length_col = next((col for col in df_filtered.columns if "length" in str(col).lower() and "carton" in str(col).lower()), None)
        width_col = next((col for col in df_filtered.columns if "width" in str(col).lower() and "carton" in str(col).lower()), None)
        height_col = next((col for col in df_filtered.columns if "height" in str(col).lower() and "carton" in str(col).lower()), None)
        weight_col = next((col for col in df_filtered.columns if "weight" in str(col).lower() and "carton" in str(col).lower()), None)

        if not length_col or not width_col or not height_col or not weight_col:
            raise ValueError("One or more carton dimension columns are missing in the supplier file.")

        duty_col = next((col for col in df_filtered.columns if "duty" in str(col).lower() and "%" in str(col)), None)
        if not duty_col:
            duty_col = next((col for col in df_filtered.columns if "duty" in str(col).lower()), None)

        if not duty_col:
            raise ValueError("Duty percentage column not found in the supplier file.")

        fob_col = next((col for col in df_filtered.columns if "fob price" in str(col).lower() and "case" in str(col).lower()), None)
        if not fob_col:
            raise ValueError("FOB price per case column not found in the supplier file.")

        packs_col = next((col for col in df_filtered.columns if "packs per case" in str(col).lower()), None)
        if not packs_col:
            raise ValueError("Column for 'Packs per case' not found in supplier file.")

        for idx, row in df_filtered.iterrows():
            if row.iloc[1] != ref:
                continue

            try:
                length = float(row[length_col])
                width = float(row[width_col])
                height = float(row[height_col])
                volume_cuft = round((length * width * height) / 1728, 2)
            except:
                volume_cuft = 0
                length = width = height = 0

            try:
                weight_carton = float(row[weight_col])
            except:
                weight_carton = 0

            try:
                packs_per_case = float(row[packs_col])
            except:
                packs_per_case = 0

            try:
                duty_pct = row[duty_col] if pd.notnull(row[duty_col]) and row[duty_col] != "N/A" else 0
                duty_pct = round(float(duty_pct), 2)

                fob_price = row[fob_col] if pd.notnull(row[fob_col]) and row[fob_col] != "N/A" else 0
                fob_price = round(float(fob_price), 2)
            except Exception as e:
                duty_pct = 0
                fob_price = 0

            try:
                total_cases_per_pallet  = TI * HI
            except:
                total_cases_per_pallet = 0

            try:
                Inbound_to_EAST = inbound_to_east_ocean_freight_amount + drayage_cost
                pallets_per_container = round(cases_fit / total_cases_per_pallet) if total_cases_per_pallet > 0 else 0
               
                duty_case = round(fob_price * (Total_duty/100), 2)
               
                ocean_per_case = round(inbound_to_east_ocean_freight_amount / cases_fit, 2) if cases_fit > 0 else 0
                inbound_per_case = round(Inbound_to_EAST / cases_fit, 2) if cases_fit > 0 else 0
                Inbound_pct_to_supplier_FOB_price = round((inbound_per_case / fob_price) * 100, 2) if fob_price > 0 else 0
               
                Artwork = round((3800) / (2 * annual_vol), 2) if annual_vol > 0 else 0
                domestic_compliance = round((500 / annual_vol), 2) if annual_vol > 0 else 0
                certification_per_case_2yrs = round(certification_cost / (2 * annual_vol), 2) if annual_vol > 0 else 0
               
                if(Warehouse_inputs == "TAL"):
                    storage_cost_per_case = round(((storage_cost_per_pallet * (storage_months - 1)) / total_cases_per_pallet), 2) if total_cases_per_pallet > 0 else 0
                    pallet_cost = round((pallet_handling_cost / total_cases_per_pallet), 2) if total_cases_per_pallet > 0 else 0
                else:
                    storage_cost_per_case = round((18.5* (storage_months) / total_cases_per_pallet), 2) if total_cases_per_pallet > 0 else 0
                    pallet_cost = round(((14+12.5+5.75) / total_cases_per_pallet), 2) if total_cases_per_pallet > 0 else 0

                Warehouse_pct_to_supplier_FOB_price = round(((pallet_cost + storage_cost_per_case) / fob_price) * 100, 2) if fob_price > 0 else 0
               
                Financing_Cost_Months = 2 + storage_months
                Overhead = round((0.05 + (0.005 * Financing_Cost_Months)) * 100, 2)
               
                No_of_cases_in_30_Pallets = 30 * total_cases_per_pallet
                No_of_cases_in_6_Pallets = 6 * total_cases_per_pallet
                No_of_cases_in_3_Pallets = 3 * total_cases_per_pallet
               
            except Exception as e:
                pallets_per_container = 0
                ocean_per_case = 0
                inbound_per_case = 0
                Inbound_pct_to_supplier_FOB_price = 0
                Artwork = 0
                domestic_compliance = 0
                certification_per_case_2yrs = 0
                pallet_cost = 0
                storage_cost_per_case = 0
                Warehouse_pct_to_supplier_FOB_price = 0
                Financing_Cost_Months = 0
                Overhead = 0
                No_of_cases_in_30_Pallets = 0
                No_of_cases_in_6_Pallets = 0
                No_of_cases_in_3_Pallets = 0
                duty_case = 0
                Total_duty = 0

            try:
                fob_price = round(pd.to_numeric(fob_price, errors='coerce'), 2) if pd.notnull(fob_price) else 0
                duty_case = round(pd.to_numeric(duty_case, errors='coerce'), 2) if pd.notnull(duty_case) else 0
                inbound_per_case = round(pd.to_numeric(inbound_per_case, errors='coerce'), 2) if pd.notnull(inbound_per_case) else 0
                Artwork = round(pd.to_numeric(Artwork, errors='coerce'), 2) if pd.notnull(Artwork) else 0
                domestic_compliance = round(pd.to_numeric(domestic_compliance, errors='coerce'), 2) if pd.notnull(domestic_compliance) else 0
                certification_per_case_2yrs = round(pd.to_numeric(certification_per_case_2yrs, errors='coerce'), 2) if pd.notnull(certification_per_case_2yrs) else 0
                pallet_cost = round(pd.to_numeric(pallet_cost, errors='coerce'), 2) if pd.notnull(pallet_cost) else 0
                storage_cost_per_case = round(pd.to_numeric(storage_cost_per_case, errors='coerce'), 2) if pd.notnull(storage_cost_per_case) else 0
           
                total_per_case = round(sum([
                    fob_price,
                    duty_case,
                    inbound_per_case,
                    Artwork,
                    domestic_compliance,
                    certification_per_case_2yrs,
                    pallet_cost,
                    storage_cost_per_case,
                    chargeback_fee
                ]), 2)

            except Exception as e:
                total_per_case = 0

            try:
                Overhead_Per_case = (Overhead / 100) * total_per_case
                Overhead_total = Overhead_Per_case * annual_vol
                Total_US_FOB_per_case = total_per_case + Overhead_Per_case
                markup_per_case = (markup/100) * Total_US_FOB_per_case if Total_US_FOB_per_case > 0 else 0
                Total_US_FOB_per_case_pickup = Total_US_FOB_per_case + markup_per_case
               
                Overhead_Per_case = round(Overhead_Per_case, 2)
                Overhead_total = round(Overhead_total, 2)
                Total_US_FOB_per_case = round(Total_US_FOB_per_case, 2)
                markup_per_case = round(markup_per_case, 2)
                Total_US_FOB_per_case_pickup = round(Total_US_FOB_per_case_pickup, 2)
               
                Total_US_FOB_Pack_pickup = round(Total_US_FOB_per_case_pickup / float(packs_per_case), 2) if packs_per_case > 0 else 0
               
                TCP_NET_Profit_Margin = (markup_per_case / Total_US_FOB_per_case_pickup) if Total_US_FOB_per_case_pickup > 0 else 0
               
                Finalized_FOB_perCase_pct_to_US_FOB_Pickup_price = (fob_price / Total_US_FOB_per_case_pickup) if Total_US_FOB_per_case_pickup > 0 else 0
                Duty_pct_to_US_FOB_Pickup_price = (duty_case / Total_US_FOB_per_case_pickup) if Total_US_FOB_per_case_pickup > 0 else 0
                Inbound_pct_to_US_FOB_Pickup_price = (inbound_per_case / Total_US_FOB_per_case_pickup) if Total_US_FOB_per_case_pickup > 0 else 0
                Artwork_pct_to_US_FOB_Pickup_price = (Artwork / Total_US_FOB_per_case_pickup) if Total_US_FOB_per_case_pickup > 0 else 0
                Domestic_Compliance_pct_to_US_FOB_Pickup_price = (domestic_compliance / Total_US_FOB_per_case_pickup) if Total_US_FOB_per_case_pickup > 0 else 0
                Certification_pct_to_US_FOB_Pickup_price = (certification_per_case_2yrs / Total_US_FOB_per_case_pickup) if Total_US_FOB_per_case_pickup > 0 else 0
                Warehouse_pct_to_US_FOB_pickup_price = ((pallet_cost + storage_cost_per_case) / Total_US_FOB_per_case_pickup) if Total_US_FOB_per_case_pickup > 0 else 0  
                Overhead_Per_case_pct_to_US_FOB_pickup_price = (Overhead_Per_case / Total_US_FOB_per_case_pickup) if Total_US_FOB_per_case_pickup > 0 else 0
               
                TOTAL_pct_components_to_US_FOB_pickup_price = round(
                    (Finalized_FOB_perCase_pct_to_US_FOB_Pickup_price +
                    Duty_pct_to_US_FOB_Pickup_price +
                    Inbound_pct_to_US_FOB_Pickup_price +
                    Artwork_pct_to_US_FOB_Pickup_price +
                    Domestic_Compliance_pct_to_US_FOB_Pickup_price +
                    Certification_pct_to_US_FOB_Pickup_price +
                    Warehouse_pct_to_US_FOB_pickup_price +
                    Overhead_Per_case_pct_to_US_FOB_pickup_price +
                    TCP_NET_Profit_Margin),2
                ) * 100

               
                Finalized_FOB_perCase_pct_to_US_FOB_Pickup_price = round(Finalized_FOB_perCase_pct_to_US_FOB_Pickup_price, 2) * 100
                Duty_pct_to_US_FOB_Pickup_price = round(Duty_pct_to_US_FOB_Pickup_price, 2) * 100
                Inbound_pct_to_US_FOB_Pickup_price = round(Inbound_pct_to_US_FOB_Pickup_price, 2) * 100
                Artwork_pct_to_US_FOB_Pickup_price = round(Artwork_pct_to_US_FOB_Pickup_price, 2) * 100
                Domestic_Compliance_pct_to_US_FOB_Pickup_price = round(Domestic_Compliance_pct_to_US_FOB_Pickup_price, 2) * 100
                Certification_pct_to_US_FOB_Pickup_price = round(Certification_pct_to_US_FOB_Pickup_price, 2) * 100
                Warehouse_pct_to_US_FOB_pickup_price = round(Warehouse_pct_to_US_FOB_pickup_price, 2) * 100
                Overhead_Per_case_pct_to_US_FOB_pickup_price = round(Overhead_Per_case_pct_to_US_FOB_pickup_price, 2) * 100
                TCP_NET_Profit_Margin = round((markup_per_case / Total_US_FOB_per_case_pickup),2) * 100 if Total_US_FOB_per_case_pickup > 0 else 0

               
                FTL_outfreight_per_case = round(FTL_Outfreight / No_of_cases_in_30_Pallets, 2) if No_of_cases_in_30_Pallets > 0 else 0
                FTL_Delivered_case_price = round(FTL_outfreight_per_case + Total_US_FOB_per_case_pickup, 2)
                FTL_Outfreight_per_FTL_Delivered = round((FTL_outfreight_per_case / FTL_Delivered_case_price) * 100, 2) if FTL_Delivered_case_price > 0 else 0
                FTL_Delivered_pack_price = round(FTL_Delivered_case_price / packs_per_case, 2) if packs_per_case > 0 else 0
               
                LTL_Outfreight_6 = LTL_Outfreight * 6
                LTL_with_Markup_15 = round(LTL_Outfreight_6 + (LTL_Outfreight_6 * 0.15), 2)
                LTL_outfreight_per_case = round(LTL_with_Markup_15 / No_of_cases_in_6_Pallets, 2) if No_of_cases_in_6_Pallets > 0 else 0
                LTL_Delivered_case_price = round(LTL_outfreight_per_case + Total_US_FOB_per_case_pickup, 2)
                LTL_Delivered_pack_price = round(LTL_Delivered_case_price / packs_per_case, 2) if packs_per_case > 0 else 0
                LTL_Outfreight_per_LTL_Delivered = round((LTL_outfreight_per_case / LTL_Delivered_case_price) * 100, 2) if LTL_Delivered_case_price > 0 else 0
               
                LTL_Outfreight_3 = LTL_Outfreight * 3
                LTL_with_Markup_15_3 = round(LTL_Outfreight_3 + (LTL_Outfreight_3 * 0.15), 2)
                LTL_outfreight_per_case_3 = round(LTL_with_Markup_15_3 / No_of_cases_in_3_Pallets, 2) if No_of_cases_in_3_Pallets > 0 else 0
                LTL_Delivered_case_price_3 = round(LTL_outfreight_per_case_3 + Total_US_FOB_per_case_pickup, 2)
                LTL_Delivered_pack_price_3 = round(LTL_Delivered_case_price_3 / packs_per_case, 2) if packs_per_case > 0 else 0
                LTL_Outfreight_per_LTL_Delivered_3 = round((LTL_outfreight_per_case_3 / LTL_Delivered_case_price_3) * 100, 2) if LTL_Delivered_case_price_3 > 0 else 0
               
                Total_Revenue_by_Case_pickup = round(annual_vol * Total_US_FOB_per_case_pickup, 2)
                Net_Total_Profit_pickup = round(annual_vol * markup_per_case, 2)
               
                Total_Revenue_by_Case_FTL = round(annual_vol * FTL_Delivered_case_price, 2)
                Net_Total_Profit_FTL = round(annual_vol * (FTL_Delivered_case_price - FTL_outfreight_per_case - Total_US_FOB_per_case), 2)
               
                Total_Revenue_by_Case_LTL_6 = round(annual_vol * LTL_Delivered_case_price, 2)
                Net_Total_Profit_LTL_6 = round(annual_vol * (LTL_Delivered_case_price - LTL_outfreight_per_case - Total_US_FOB_per_case), 2)
               
                Total_Revenue_by_Case_LTL_3 = round(annual_vol * LTL_Delivered_case_price_3, 2)
                Net_Total_Profit_LTL_3 = round(annual_vol * (LTL_Delivered_case_price_3 - LTL_outfreight_per_case_3 - Total_US_FOB_per_case), 2)
               
                Customer_Margin_pct_based_on_Pickup = round(((selling_price - Total_US_FOB_Pack_pickup) / selling_price) * 100, 2) if selling_price > 0 else 0
                Customer_Margin_pct_based_on_FTL = round(((selling_price - FTL_Delivered_pack_price) / selling_price) * 100, 2) if selling_price > 0 else 0
                Customer_Margin_pct_based_on_6_pallet_LTL = round(((selling_price - LTL_Delivered_pack_price) / selling_price) * 100, 2) if selling_price > 0 else 0
                Customer_Margin_pct_based_on_3_pallet_LTL = round(((selling_price - LTL_Delivered_pack_price_3) / selling_price) * 100, 2) if selling_price > 0 else 0
               
            except Exception as e:
                Overhead_Per_case = 0
                Overhead_total = 0
                Total_US_FOB_per_case = 0
                markup_per_case = 0
                Total_US_FOB_per_case_pickup = 0
                Total_US_FOB_Pack_pickup = 0
                TCP_NET_Profit_Margin = 0
                Finalized_FOB_perCase_pct_to_US_FOB_Pickup_price = 0
                Duty_pct_to_US_FOB_Pickup_price = 0
                Inbound_pct_to_US_FOB_Pickup_price = 0
                Artwork_pct_to_US_FOB_Pickup_price = 0
                Domestic_Compliance_pct_to_US_FOB_Pickup_price = 0
                Certification_pct_to_US_FOB_Pickup_price = 0
                Warehouse_pct_to_US_FOB_pickup_price = 0
                Overhead_Per_case_pct_to_US_FOB_pickup_price = 0
                FTL_outfreight_per_case = 0
                FTL_Delivered_case_price = 0
                FTL_Outfreight_per_FTL_Delivered = 0
                FTL_Delivered_pack_price = 0
                LTL_Outfreight_6 = 0
                LTL_with_Markup_15 = 0
                LTL_outfreight_per_case = 0
                LTL_Delivered_case_price = 0
                LTL_Delivered_pack_price = 0
                LTL_Outfreight_3 = 0
                LTL_with_Markup_15_3 = 0
                LTL_outfreight_per_case_3 = 0
                LTL_Delivered_case_price_3 = 0
                LTL_Delivered_pack_price_3 = 0
                LTL_Outfreight_per_LTL_Delivered_3 = 0
                Total_Revenue_by_Case_pickup = 0
                Net_Total_Profit_pickup = 0
                Total_Revenue_by_Case_FTL = 0
                Net_Total_Profit_FTL = 0
                Total_Revenue_by_Case_LTL_6 = 0
                Net_Total_Profit_LTL_6 = 0
                Total_Revenue_by_Case_LTL_3 = 0
                Net_Total_Profit_LTL_3 = 0
                Customer_Margin_pct_based_on_Pickup = 0
                Customer_Margin_pct_based_on_FTL = 0
                Customer_Margin_pct_based_on_6_pallet_LTL = 0
                Customer_Margin_pct_based_on_3_pallet_LTL = 0
                TOTAL_pct_components_to_US_FOB_pickup_price = 0

           
            results[ref] = {
                "Finalized Supplier": min_supplier,
                "No. of cases fit into container": round(cases_fit),
                "No. of pallets can fit into a container": round(pallets_per_container),
                "Cases per pallet": total_cases_per_pallet,
                "Finalized FOB/case": "$"+str(round(fob_price, 2)),
                "Finalized FOB/ Case % to US FOB Pickup price": str(round(Finalized_FOB_perCase_pct_to_US_FOB_Pickup_price, 2)) + "%",
                "Finalized Duty% based on finalized supplier": str(round(duty_pct, 2)) + "%",
                "General duty%": str(round(General_duty_pct, 2)) + "%",
                "General duty $": "$" + str(round(General_duty_price, 2)),
                "Reciprocal %": str(round(Reciprocal_pct, 2)) + "%",
                # "Reciprocal Per Case": round(Reciprocal_Per_Case, 2),
                "Section 301 China %": str(Section_301_china) + "%",
                "Section 232 duty 25%": str(round(Section_232_duty, 2)) + "%",
                "Tariff %": str(Tariff) + "%",
                "Total duty %": str(round(Total_duty, 2)) + "%",
                "Duty/Case": "$"+str(round(duty_case, 2)),
                "Duty % to US FOB Pickup price": str(round(Duty_pct_to_US_FOB_Pickup_price, 2)) + "%",
                "Inbound to EAST Ocean freight": "$"+str(round(inbound_to_east_ocean_freight_amount, 2)),
                "Ocean per case": "$"+str(round(ocean_per_case, 2)),
                "Inbound to EAST (Ocean = $ 5500, Drayage = $1300, Unloading= $0/container)": "$"+str(round(Inbound_to_EAST, 2)),
                "Inbound/case": "$"+str(round(inbound_per_case, 2)),
                "Inbound % to supplier FOB price": str(round(Inbound_pct_to_supplier_FOB_price, 2)) + "%",
                "Inbound % to US FOB Pickup price": str(round(Inbound_pct_to_US_FOB_Pickup_price, 2)) + "%",
                "Artwork ($/SKU) =3800 ; Testing/QA=0 Per Case Calc ($)": "$"+str(round(Artwork, 2)),
                "Artwork % to US FOB Pickup price": str(round(Artwork_pct_to_US_FOB_Pickup_price, 2)) + "%",
                "Domestic Compliance (qulaity assurance) = $500": "$"+str(round(domestic_compliance, 2)),
                "Domestic Compliance % to US FOB Pickup price": str(round(Domestic_Compliance_pct_to_US_FOB_Pickup_price, 2)) + "%",
                "Certification = $1000/facility (GFSI- food safety) per case for 2 Years": "$"+str(round(certification_per_case_2yrs, 2)),
                "Certification % to US FOB Pickup price": str(round(Certification_pct_to_US_FOB_Pickup_price, 2)) + "%",
                "Palletizing cost per case (TAL)" if Warehouse_inputs == "TAL" else "Palletizing cost per case (ELM)" : "$"+str(round(pallet_cost, 2)),
                "Storage Months": round(storage_months, 2),
                "Storage cost per case (TAL)" if Warehouse_inputs == "TAL" else "Storage cost per case (ELM)": "$"+str(round(storage_cost_per_case, 2)),
                "Warehouse % to supplier FOB price": str(round(Warehouse_pct_to_supplier_FOB_price, 2)) + "%",
                "Warehouse % to US FOB pickup price": str(round(Warehouse_pct_to_US_FOB_pickup_price, 2)) + "%",
                "Chargeback Fee": "$" + str(round(chargeback_fee, 2)),
                "Total/Case": "$"+str(round(total_per_case, 2)),
                "Financing Cost Months (1 mth Prod + 1 mth ship + Storage Months)": round(Financing_Cost_Months, 2),
                "Overhead Sales=1%, Daymon=0%; Supplier service fees- 0%; Payment Term (2%30 Net 30)= 3%; Delivery Issues= 1%; Reclamation= 0%, Financing Cost=0.5 % * Months": str(round(Overhead, 2)) + "%",
                "Overhead Per case": "$"+str(round(Overhead_Per_case, 2)),
                "Overhead Per case % to US FOB pickup price": str(round(Overhead_Per_case_pct_to_US_FOB_pickup_price, 2)) + "%",
                "Overhead Total ($)": "$"+str(round(Overhead_total, 2)),
                "Total US FOB/case": "$"+str(round(Total_US_FOB_per_case, 2)),
                "Markup %": str(round(markup, 2)) + "%",
                "Markup/Case": "$"+str(round(markup_per_case, 2)),
                "Total US FOB /Case pickup": "$"+str(round(Total_US_FOB_per_case_pickup, 2)),
                "Total US FOB /Pack pickup": "$"+str(round(Total_US_FOB_Pack_pickup, 2)),
                "TCP NET Profit Margin": str(round(TCP_NET_Profit_Margin, 2)) + "%",
                "TOTAL % components to US FOB pickup price": str(round(TOTAL_pct_components_to_US_FOB_pickup_price, 2)) + "%",
                "FTL Outfreight": "$" + str(round(FTL_Outfreight, 2)),
                "No. of cases in 30 Pallets": round(No_of_cases_in_30_Pallets, 2),
                "FTL outfreight/case": "$" + str(round(FTL_outfreight_per_case, 2)),
                "FTL Outfreight/FTL Delivered": str(round(FTL_Outfreight_per_FTL_Delivered, 2)) + "%",
                "FTL Delivered pack price": "$" + str(round(FTL_Delivered_pack_price, 2)),
                "FTL Delivered case price": "$" + str(round(FTL_Delivered_case_price, 2)),
                "LTL Outfreight": "$" + str(round(LTL_Outfreight, 2)),
                "LTL Outfreight 6 pallets": "$" + str(round(LTL_Outfreight_6, 2)),
                "LTL with Markup 15%": "$" + str(round(LTL_with_Markup_15, 2)),
                "No. of cases in 6 Pallets": round(No_of_cases_in_6_Pallets, 2),
                "LTL outfreight/case": "$" + str(round(LTL_outfreight_per_case, 2)),
                "LTL Outfreight/LTL Delivered": str(round(LTL_Outfreight_per_LTL_Delivered, 2)) + "%",
                "LTL Delivered pack price": "$" + str(round(LTL_Delivered_pack_price, 2)),
                "LTL Delivered case price": "$" + str(round(LTL_Delivered_case_price, 2)),
                "LTL Outfreight 3 pallets": "$" + str(round(LTL_Outfreight_3, 2)),
                "LTL with Markup 15% for 3 pallets": "$" + str(round(LTL_with_Markup_15_3, 2)),
                "No. of cases in 3 Pallets": round(No_of_cases_in_3_Pallets, 2),
                "LTL outfreight/case for 3 pallets": "$" + str(round(LTL_outfreight_per_case_3, 2)),
                "LTL Outfreight/LTL Delivered for 3 pallets": str(round(LTL_Outfreight_per_LTL_Delivered_3, 2)) + "%",
                "LTL Delivered PACK price for 3 pallets": "$" + str(round(LTL_Delivered_pack_price_3, 2)),
                "LTL Delivered case price for 3 pallets": "$" + str(round(LTL_Delivered_case_price_3, 2)),
                "Annual Volume Cases QTY": annual_vol,
                "Total Revenue by Case (Pickup)": "$" + str(round(Total_Revenue_by_Case_pickup, 2)),
                "Net Total Profit (Pickup)": "$" + str(round(Net_Total_Profit_pickup, 2)),
                "Total Revenue by Case (FTL)": "$" + str(round(Total_Revenue_by_Case_FTL, 2)),
                "Net Total Profit (FTL)": "$" + str(round(Net_Total_Profit_FTL, 2)),
                "Total Revenue by Case (LTL) 6 pallets": "$" + str(round(Total_Revenue_by_Case_LTL_6, 2)),
                "Net Total Profit (LTL) 6 pallets": "$" + str(round(Net_Total_Profit_LTL_6, 2)),
                "Total Revenue by Case (LTL) 3 pallets": "$" + str(round(Total_Revenue_by_Case_LTL_3, 2)),
                "Net Total Profit (LTL) 3 pallets": "$" + str(round(Net_Total_Profit_LTL_3, 2)),
                "Weblink": "",
                "Selling price": "$" + str(round(selling_price, 2)),
                "Customer Margin % based on Pickup": str(round(Customer_Margin_pct_based_on_Pickup, 2)) + "%",
                "Customer Margin % based on FTL": str(round(Customer_Margin_pct_based_on_FTL, 2)) + "%",
                "Customer Margin % based on 6 pallet LTL": str(round(Customer_Margin_pct_based_on_6_pallet_LTL, 2)) + "%",
                "Customer Margin % based on 3 pallet LTL": str(round(Customer_Margin_pct_based_on_3_pallet_LTL, 2)) + "%",
                "NBE selling price": "",
            }

    return results

# ----------------- CREATE FORMATTING DATAFRAME ----------------- #
def create_formatting_dataframe():
    headers = [
        "Min FOB/Case",
        "Min Supplier",
        "Finalized Supplier",
        "COO of Finalized Supplier",
        "No. of cases fit into container",
        "No. of pallets can fit into a container",
        "Cases per pallet",
        "Finalized FOB/case",
        "Finalized FOB/ Case % to US FOB Pickup price",
        "Finalized Duty% based on finalized supplier",
        "General duty%",
        "General duty $",
        "Reciprocal %",
        # "Reciprocal Per Case",
        "Section 301 China %",
        "Section 232 duty 25%",
        "Tariff %",
        "Total duty %",
        "Duty/Case",
        "Duty % to US FOB Pickup price",
        "Inbound to EAST Ocean freight",
        "Ocean per case",
        "Inbound to EAST (Ocean = $ 5500, Drayage = $1300, Unloading= $0/container)",
        "Inbound/case",
        "Inbound % to supplier FOB price",
        "Inbound % to US FOB Pickup price",
        "Artwork ($/SKU) =3800 ; Testing/QA=0 Per Case Calc ($)",
        "Artwork % to US FOB Pickup price",
        "Domestic Compliance (qulaity assurance) = $500",
        "Domestic Compliance % to US FOB Pickup price",
        "Chargeback Fee = $",
        "Certification = $1000/facility (GFSI- food safety) per case for 2 Years",
        "Certification % to US FOB Pickup price",
        "Palletizing cost per case",
        "Storage Months",
        "Storage cost per case",
        "Warehouse % to supplier FOB price",
        "Warehouse % to US FOB pickup price",
        "Total/Case",
        "Financing Cost Months (1 mth Prod + 1 mth ship + Storage Months)",
        "Overhead Sales=1%, Daymon=0%; Supplier service fees- 0%; Payment Term (2%30 Net 30)= 3%; Delivery Issues= 1%; Reclamation= 0%, Financing Cost=0.5 % * Months",
        "Overhead Per case",
        "Overhead Per case % to US FOB pickup price",
        "Overhead Total ($)",
        "Total US FOB/case",
        "Markup %",
        "Markup/Case",
        "Total US FOB /Case pickup",
        "Total US FOB /Pack pickup",
        "TCP NET Profit Margin",
        "TOTAL % components to US FOB pickup price",
        "FTL Outfreight",
        "No. of cases in 30 Pallets",
        "FTL outfreight/case",
        "FTL Outfreight/FTL Delivered",
        "FTL Delivered pack price",
        "FTL Delivered case price",
        "LTL Outfreight 6 pallets",
        "LTL with Markup 15%",
        "No. of cases in 6 Pallets",
        "LTL outfreight/case",
        "LTL Outfreight/LTL Delivered",
        "LTL Delivered pack price",
        "LTL Delivered case price",
        "LTL Outfreight 3 pallets",
        "LTL with Markup 15% for 3 pallets",
        "No. of cases in 3 Pallets",
        "LTL outfreight/case for 3 pallets",
        "LTL Outfreight/LTL Delivered for 3 pallets",
        "LTL Delivered PACK price for 3 pallets",
        "LTL Delivered case price for 3 pallets",
        "Annual Volume Cases QTY",
        "Total Revenue by Case (Pickup)",
        "Net Total Profit (Pickup)",
        "Total Revenue by Case (FTL)",
        "Net Total Profit (FTL)",
        "Total Revenue by Case (LTL) 6 pallets",
        "Net Total Profit (LTL) 6 pallets",
        "Total Revenue by Case (LTL) 3 pallets",
        "Net Total Profit (LTL) 3 pallets",
        "Weblink",
        "Selling price",
        "Customer Margin % based on Pickup",
        "Customer Margin % based on FTL",
        "Customer Margin % based on 6 pallet LTL",
        "Customer Margin % based on 3 pallet LTL",
        "NBE selling price"
    ]
    return pd.DataFrame({"Header": headers})


# ----------------- CREATE WORKBOOK WITH TCP INFO ----------------- #
def create_workbook(all_refs, file_data, title_rows, file_names, calculation_results, tcp_inputs, comparison_results, tcp_headers=None, tcp_ref_idx=None, tcp_data = None):

    pallet_length = tcp_inputs.get('pallet_length', 40)
    pallet_width = tcp_inputs.get('pallet_width', 48)
    max_pallet_height = tcp_inputs.get('max_pallet_height', 85)
    max_pallet_weight = tcp_inputs.get('max_pallet_weight', 1200)
    Warehouse_inputs = tcp_inputs["Warehouse_inputs"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Merged Supplier Analysis"

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
   
    supplier_colors = [
        "FF98d454",  # Light green
    ]
   
    data_row_color = "E8F5E8"
    pricing_color = "FFfffc04"
    tcp_color = "B7DEE8"

    current_row = 1
    col_offset = 1

    # ------------------ TCP Number Column (Single Column) ------------------ #
    tcp_number_col = col_offset
    tcp_header = ws.cell(row=current_row+1, column=tcp_number_col, value="TCP Number")
    tcp_header.font = Font(bold=True, size=12)
    tcp_header.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    tcp_header.border = thin_border
    tcp_header.alignment = Alignment(horizontal="center", vertical="center")
   
    col_offset += 1

    # ------------------ TCP Info Section ------------------ #
    tcp_col_start = col_offset
    tcp_data_headers = []
   
    if tcp_headers and tcp_data:
        for i, header in enumerate(tcp_headers):
            if i != tcp_ref_idx:
                tcp_data_headers.append((i, header))
       
        num_tcp_cols = len(tcp_data_headers)
       
        if num_tcp_cols > 0:
            tcp_merge_end = col_offset + num_tcp_cols - 1
           
            ws.merge_cells(start_row=current_row, start_column=col_offset,
                          end_row=current_row, end_column=tcp_merge_end)
           
            tcp_title_cell = ws.cell(row=current_row, column=col_offset, value="TCP Information")
            tcp_title_cell.font = Font(bold=True, size=16, color="000000")
            tcp_title_cell.fill = PatternFill(start_color=tcp_color, end_color=tcp_color, fill_type="solid")
            tcp_title_cell.alignment = Alignment(horizontal="center", vertical="center")
            tcp_title_cell.border = thin_border
           
            current_col = col_offset
            for orig_idx, header in tcp_data_headers:
                # header is a plain string, not a dict
                cell = ws.cell(row=current_row + 1, column=current_col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                current_col += 1
           
            col_offset = current_col + 1

    # ------------------ Supplier Sections ------------------ #
    column_positions = []

    for idx, ((title_row, _), (headers, ref_idx, _)) in enumerate(zip(title_rows, file_data)):
        supplier_name = file_names[idx].replace(".xlsx", "")
        supplier_color = supplier_colors[idx % len(supplier_colors)]
       
        data_headers = []
        for i, header in enumerate(headers):
            header_value = (str(header["value"]).strip().lower() if header["value"] else "")
            # common variants that mean "reference/TCP"
            ref_markers = ["tcp number", "reference number", "sr.no", "no.", "no"]

            if i != ref_idx and header_value not in ref_markers:
                data_headers.append(header)
           
        num_data_cols = len(data_headers)
        merge_end = col_offset + num_data_cols - 1
       
        ws.merge_cells(start_row=current_row, start_column=col_offset,
                      end_row=current_row, end_column=merge_end)
       
        cell = ws.cell(row=current_row, column=col_offset, value=supplier_name)
        cell.font = Font(bold=True, size=16, color="000000")
        cell.fill = PatternFill(start_color=supplier_color, end_color=supplier_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
       
        header_start_col = col_offset
        current_col = col_offset
       
        for header in data_headers:
            cell = ws.cell(row=current_row + 1, column=current_col, value=header["value"])
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            current_col += 1
       
        column_positions.append((header_start_col, current_col - 1, ref_idx, supplier_color, data_headers))
       
        col_offset = current_col
        if idx < len(file_data) - 1:
            col_offset += 1
 
    supplier_dimension_cols = {}
    for (start_col, end_col, ref_col_idx, supplier_color, data_headers), fname in zip(column_positions, file_names):
        supplier_name = fname.replace(".xlsx", "").strip().lower()
        supplier_dimension_cols[supplier_name] = {}
       
        for idx, header in enumerate(data_headers):
            if header["value"]:
                header_lower = str(header["value"]).strip().lower()
                col_letter = get_column_letter(start_col + idx)
               
                if "carton width" in header_lower or "width" in header_lower:
                    supplier_dimension_cols[supplier_name]["width"] = col_letter
                elif "carton length" in header_lower or "length" in header_lower:
                    supplier_dimension_cols[supplier_name]["length"] = col_letter
                elif "carton height" in header_lower or "height" in header_lower:
                    supplier_dimension_cols[supplier_name]["height"] = col_letter
                elif "carton weight" in header_lower or "weight" in header_lower:
                    supplier_dimension_cols[supplier_name]["weight"] = col_letter
                elif "packs per case" in header_lower:
                    supplier_dimension_cols[supplier_name]["packs_per_case"] = col_letter

    formatting_df = create_formatting_dataframe()

    if Warehouse_inputs == "TAL":
        formatting_df["Header"] = formatting_df["Header"].replace({
            "Palletizing cost per case": "Palletizing cost per case (TAL)",
            "Storage cost per case": "Storage cost per case (TAL)"
        })
    elif Warehouse_inputs == "ELM":
        formatting_df["Header"] = formatting_df["Header"].replace({
            "Palletizing cost per case": "Palletizing cost per case (ELM)",
            "Storage cost per case": "Storage cost per case (ELM)"
        })

    calculated_col_start = col_offset + 1
   
    pricing_merge_end = calculated_col_start + len(formatting_df) - 1
    ws.merge_cells(start_row=current_row, start_column=calculated_col_start,
                  end_row=current_row, end_column=pricing_merge_end)
   
    pricing_cell = ws.cell(row=current_row, column=calculated_col_start, value="Pricing")
    pricing_cell.font = Font(bold=True, size=16, color="000000")
    pricing_cell.fill = PatternFill(start_color=pricing_color, end_color=pricing_color, fill_type="solid")
    pricing_cell.alignment = Alignment(horizontal="center", vertical="center")
    pricing_cell.border = thin_border

    calc_header_row = current_row + 1
    for idx, row in formatting_df.iterrows():
        cell = ws.cell(row=calc_header_row, column=calculated_col_start + idx, value=row["Header"])
        cell.fill = PatternFill(start_color="FF98d454", end_color="FF98d454", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    formula_text_row = calc_header_row + 1
    formula_row = formula_text_row + 1

    header_to_col = {row["Header"]: calculated_col_start + idx for idx, row in formatting_df.iterrows()}

    finalized_supplier_col = header_to_col.get("Finalized Supplier")
    if finalized_supplier_col is None:
        raise ValueError("Finalized Supplier column not found in Pricing headers")

    finalized_supplier_col_letter = get_column_letter(finalized_supplier_col)

    def ref_data_val(hdr):
        c = header_to_col.get(hdr)
        if c:
            return f"{get_column_letter(c)}{formula_row}"
        else:
            return None

    def ref_data_isnum(hdr):
        c = header_to_col.get(hdr)
        if c:
            col_letter = get_column_letter(c)
            return f"ISNUMBER({col_letter}{formula_row})"
        else:
            return None

    def isnumber_and(*hdrs):
        checks = [ref_data_isnum(h) for h in hdrs if ref_data_isnum(h)]
        if checks:
            return "AND(" + ",".join(checks) + ")"
        else:
            return "TRUE"

    formula_descriptions = {
        "Min FOB/Case": "Minimum FOB Price per Case among suppliers",
        "Min Supplier": "Supplier with Minimum FOB Price",
        "Finalized Supplier": "",
        "No. of cases fit into container": "",
        "No. of pallets can fit into a container": "No. of cases fit into container/ Cases per pallet",
        "Cases per pallet": "TI x HI (calculated using carton dimensions and pallet constraints)",
        "Finalized FOB/case": "Supplier FOB Price per Case",
        "Finalized FOB/ Case % to US FOB Pickup price": "Finalized FOB per case / Total US FOB per Case pickup",
        "Finalized Duty% based on finalized supplier": "",
        "General duty%": "",
        "General duty $": "",
        "Reciprocal %": "",
        # "Reciprocal Per Case": "",
        "Section 301 China %": "",
        "Section 232 duty 25%": "",
        "Tariff %":"",
        "Total duty %": "General duty% + Reciprocal % + Section 301 China % +Section 232 duty % + Tariff %",
        "Duty/Case": "Finalized FOB per case * Total duty %",
        "Duty % to US FOB Pickup price": "(Duty per Case / Total US FOB per Case pickup) * 100",
        "Inbound to EAST Ocean freight": "$5500",
        "Ocean per case": "Inbound Ocean Freight / No. of cases fit into container",
        "Inbound to EAST (Ocean = $ 5500, Drayage = $1300, Unloading= $0/container)": "Ocean Freight + Drayage + Unloading",
        "Inbound/case": "Inbound to EAST / No. of cases fit into container",
        "Inbound % to supplier FOB price": "(Inbound per case / Finalized FOB per case) * 100",
        "Inbound % to US FOB Pickup price": "(Inbound per case / Total US FOB per Case pickup) * 100",
        "Artwork ($/SKU) =3800 ; Testing/QA=0 Per Case Calc ($)": "Artwork Cost / (2 * Annual Volume Cases QTY)",
        "Artwork % to US FOB Pickup price": "(Artwork / Total US FOB per Case pickup) * 100",
        "Domestic Compliance (qulaity assurance) = $500": "Domestic Compliance Cost / Annual Volume Cases QTY",
        "Domestic Compliance % to US FOB Pickup price": "(Domestic Compliance / Total US FOB per Case pickup) * 100",
        "Chargeback Fee = $": "Chargeback Fee (input)",
        "Certification = $1000/facility (GFSI- food safety) per case for 2 Years": "Certification Cost / (2 * Annual Volume Cases QTY)",
        "Certification % to US FOB Pickup price": "(Certification / Total US FOB per Case pickup) * 100",
        "Palletizing cost per case (TAL)" if Warehouse_inputs == "TAL" else "Palletizing cost per case (ELM)": "Pallet Handling Cost / Cases per pallet" if Warehouse_inputs == "TAL" else "(14.5+12.5+5.75)/Cases per pallet",
        "Storage Months": "Storage Months (input)",
        "Storage cost per case (TAL)" if Warehouse_inputs == "TAL" else "Storage cost per case (ELM)": "(28 * (Storage Months - 1)) / Cases per pallet" if Warehouse_inputs == "TAL" else "(18.5 * (Storage Months)) / Cases per pallet",
        "Warehouse % to supplier FOB price": "((Palletizing cost + Storage cost) / Finalized FOB per case) * 100",
        "Warehouse % to US FOB pickup price": "((Palletizing cost + Storage cost) / Total US FOB per Case pickup) * 100",
        "Total/Case": "Sum of all cost components per case",
        "Financing Cost Months (1 mth Prod + 1 mth ship + Storage Months)": "2 + Storage Months",
        "Overhead Sales=1%, Daymon=0%; Supplier service fees- 0%; Payment Term (2%30 Net 30)= 3%; Delivery Issues= 1%; Reclamation= 0%, Financing Cost=0.5 % * Months": "5% + 0.5% * Financing Cost Months",
        "Overhead Per case": "Overhead %  * Total per Case",
        "Overhead Per case % to US FOB pickup price": "(Overhead Per case / Total US FOB per Case pickup) * 100",
        "Overhead Total ($)": "Overhead Per case * Annual Volume Cases QTY",
        "Total US FOB/case": "Total per Case + Overhead Per case",
        "Markup %": "Markup % (input)",
        "Markup/Case": "(Markup % / 100) * Total US FOB per Case",
        "Total US FOB /Case pickup": "Total US FOB per case + Markup per Case",
        "Total US FOB /Pack pickup": "Total US FOB per Case pickup / Packs per case",
        "TCP NET Profit Margin": "(Markup per Case / Total US FOB per Case pickup) * 100",
        "TOTAL % components to US FOB pickup price": "Sum of all % components to US FOB pickup price",
        "FTL Outfreight": "FTL Outfreight (input)",
        "No. of cases in 30 Pallets": "30 * Cases per pallet",
        "FTL outfreight/case": "FTL Outfreight / No. of cases in 30 Pallets",
        "FTL Outfreight/FTL Delivered": "(FTL outfreight per case / FTL Delivered case price) * 100",
        "FTL Delivered pack price": "FTL Delivered case price / Packs per case",
        "FTL Delivered case price": "FTL outfreight per case + Total US FOB per Case pickup",
        "LTL Outfreight 6 pallets": "LTL Outfreight (input) * 6",
        "LTL with Markup 15%": "LTL Outfreight 6 pallets + 15%",
        "No. of cases in 6 Pallets": "6 * Cases per pallet",
        "LTL outfreight/case": "LTL with Markup 15% / No. of cases in 6 Pallets",
        "LTL Outfreight/LTL Delivered": "(LTL outfreight per case / LTL Delivered case price) * 100",
        "LTL Delivered pack price": "LTL Delivered case price / Packs per case",
        "LTL Delivered case price": "LTL outfreight per case + Total US FOB per Case pickup",
        "LTL Outfreight 3 pallets": "LTL Outfreight (input) * 3",
        "LTL with Markup 15% for 3 pallets": "LTL Outfreight 3 pallets + 15%",
        "No. of cases in 3 Pallets": "3 * Cases per pallet",
        "LTL outfreight/case for 3 pallets": "LTL with Markup 15% for 3 pallets / No. of cases in 3 Pallets",
        "LTL Outfreight/LTL Delivered for 3 pallets": "(LTL outfreight per case for 3 pallets / LTL Delivered case price for 3 pallets) * 100",
        "LTL Delivered PACK price for 3 pallets": "LTL Delivered case price for 3 pallets / Packs per case",
        "LTL Delivered case price for 3 pallets": "LTL outfreight per case for 3 pallets + Total US FOB per Case pickup",
        "Annual Volume Cases QTY": "",
        "Total Revenue by Case (Pickup)": "Annual Volume Cases QTY * Total US FOB per Case pickup",
        "Net Total Profit (Pickup)": "Annual Volume Cases QTY * Markup per Case",
        "Total Revenue by Case (FTL)": "Annual Volume Cases QTY * FTL Delivered case price",
        "Net Total Profit (FTL)": "Annual Volume Cases QTY * (FTL Delivered case price - FTL outfreight per case - Total US FOB per Case)",
        "Total Revenue by Case (LTL) 6 pallets": "Annual Volume Cases QTY * LTL Delivered case price",
        "Net Total Profit (LTL) 6 pallets": "Annual Volume Cases QTY * (LTL Delivered case price - LTL outfreight per case - Total US FOB per Case)",
        "Total Revenue by Case (LTL) 3 pallets": "Annual Volume Cases QTY * LTL Delivered case price for 3 pallets",
        "Net Total Profit (LTL) 3 pallets": "Annual Volume Cases QTY * (LTL Delivered case price for 3 pallets - LTL outfreight per case for 3 pallets - Total US FOB per Case)",
        "Weblink": "N/A",
        "Selling price": "Selling Price (input)",
        "Customer Margin % based on Pickup": "((Selling price - Total US FOB /Pack pickup) / Selling price) * 100",
        "Customer Margin % based on FTL": "((Selling price - FTL Delivered pack price) / Selling price) * 100",
        "Customer Margin % based on 6 pallet LTL": "((Selling price - LTL Delivered pack price) / Selling price) * 100",
        "Customer Margin % based on 3 pallet LTL": "((Selling price - LTL Delivered PACK price for 3 pallets) / Selling price) * 100",
        "NBE selling price": "N/A"
    }

    for idx, row_format in formatting_df.iterrows():
        header = row_format["Header"]
        cell_text = ws.cell(row=formula_text_row, column=calculated_col_start + idx)
        formula_desc = formula_descriptions.get(header, "")
        cell_text.value = formula_desc if formula_desc else ""
        cell_text.font = Font(size=12, bold=True)
        cell_text.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_text.fill = PatternFill(start_color="D4D4D4", end_color="D4D4D4", fill_type="solid")
        cell_text.border = thin_border

    # Find TI and HI columns from TCP section
    ti_col = None
    hi_col = None
    if tcp_data_headers:
        for idx, (orig_idx, header) in enumerate(tcp_data_headers):
            col_header_val = str(header).strip().upper() if header else ""
            if col_header_val == "TI":
                ti_col = get_column_letter(tcp_col_start + idx)
            elif col_header_val == "HI":
                hi_col = get_column_letter(tcp_col_start + idx)
    def header_exists(h):
        """Return True if header exists in current worksheet header map."""
        return h in header_to_col

    def ref_one_of(*candidates):
        """
        Return the first available cell reference among multiple candidate headers.
        Example: ref_one_of("Chargeback Fee", "Chargeback Fee = $")
        """
        for h in candidates:
            c = header_to_col.get(h)
            if c:
                return f"{get_column_letter(c)}{formula_row}"
        return None

    # Dynamic header naming for warehouse-specific columns
    PALLET_HDR = "Palletizing cost per case (TAL)" if Warehouse_inputs.upper() == "TAL" else "Palletizing cost per case (ELM)"
    STORAGE_HDR = "Storage cost per case (TAL)" if Warehouse_inputs.upper() == "TAL" else "Storage cost per case (ELM)"

    # -------------------- APPLY FORMULAS --------------------
    for idx, row_format in formatting_df.iterrows():
        header = row_format["Header"]
        cell = ws.cell(row=formula_row, column=calculated_col_start + idx)

        # ---------- CASES PER PALLET ----------
        if header == "Cases per pallet":
            if ti_col and hi_col:
                cell.value = f"=ROUND(${ti_col}{formula_row}*${hi_col}{formula_row},2)"
            else:
                cell.value = "N/A"

        # ---------- FINALIZED FOB % OF US FOB PICKUP ----------
        elif header == "Finalized FOB/ Case % to US FOB Pickup price":
            fob = ref_data_val("Finalized FOB/case")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Finalized FOB/case", "Total US FOB /Case pickup")
            cell.value = (
                f"=IF({check},ROUND(({fob}/{total_us_fob_pickup})*100,2),0)"
                if (fob and total_us_fob_pickup)
                else "N/A"
            )

        # ---------- TOTAL DUTY % ----------
        elif header == "Total duty %":
            parts = ["General duty%", "Reciprocal %", "Section 301 China %", "Section 232 duty 25%", "Tariff %"]
            refs = [ref_data_val(p) for p in parts if ref_data_val(p)]
            check = isnumber_and(*parts)
            cell.value = (
                f"=IF({check},ROUND(" + "+".join(refs) + ",2),0)"
                if refs
                else "N/A"
            )

        # ---------- DUTY/CASE ----------
        elif header == "Duty/Case":
            fob = ref_data_val("Finalized FOB/case")
            duty = ref_data_val("Total duty %")
            check = isnumber_and("Finalized FOB/case", "Total duty %")
            cell.value = (
                f"=IF({check},ROUND({fob}*{duty}/100,2),0)"
                if (fob and duty)
                else "N/A"
            )

        # ---------- DUTY % TO US FOB PICKUP ----------
        elif header == "Duty % to US FOB Pickup price":
            duty_case = ref_data_val("Duty/Case")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Duty/Case", "Total US FOB /Case pickup")
            cell.value = (
                f"=IF({check},ROUND(({duty_case}/{total_us_fob_pickup})*100,2),0)"
                if (duty_case and total_us_fob_pickup)
                else "N/A"
            )

        # ---------- INBOUND OCEAN + DRAYAGE + UNLOADING ----------
        elif header == "Inbound to EAST (Ocean = $ 5500, Drayage = $1300, Unloading= $0/container)":
            ocean = ref_data_val("Inbound to EAST Ocean freight")
            check = isnumber_and("Inbound to EAST Ocean freight")
            cell.value = (
                f"=IF({check},ROUND({ocean}+1300+0,2),0)" if ocean else "N/A"
            )

        # ---------- OCEAN PER CASE ----------
        elif header == "Ocean per case":
            inbound_ocean = ref_data_val("Inbound to EAST Ocean freight")
            total_cases = ref_data_val("No. of cases fit into container")
            check = isnumber_and("Inbound to EAST Ocean freight", "No. of cases fit into container")
            cell.value = (
                f"=IF({check},ROUND({inbound_ocean}/{total_cases},2),0)"
                if (inbound_ocean and total_cases)
                else "N/A"
            )

        # ---------- INBOUND/CASE ----------
        elif header == "Inbound/case":
            inbound_total = ref_data_val("Inbound to EAST (Ocean = $ 5500, Drayage = $1300, Unloading= $0/container)")
            total_cases = ref_data_val("No. of cases fit into container")
            check = isnumber_and(
                "Inbound to EAST (Ocean = $ 5500, Drayage = $1300, Unloading= $0/container)",
                "No. of cases fit into container",
            )
            cell.value = (
                f"=IF({check},ROUND({inbound_total}/{total_cases},2),0)"
                if (inbound_total and total_cases)
                else "N/A"
            )

        # ---------- INBOUND % TO SUPPLIER FOB ----------
        elif header == "Inbound % to supplier FOB price":
            inbound_case = ref_data_val("Inbound/case")
            fob = ref_data_val("Finalized FOB/case")
            check = isnumber_and("Inbound/case", "Finalized FOB/case")
            cell.value = (
                f"=IF({check},ROUND(({inbound_case}/{fob})*100,2),0)"
                if (inbound_case and fob)
                else "N/A"
            )

        # ---------- INBOUND % TO US FOB PICKUP ----------
        elif header == "Inbound % to US FOB Pickup price":
            inbound_case = ref_data_val("Inbound/case")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Inbound/case", "Total US FOB /Case pickup")
            cell.value = (
                f"=IF({check},ROUND(({inbound_case}/{total_us_fob_pickup})*100,2),0)"
                if (inbound_case and total_us_fob_pickup)
                else "N/A"
            )
                # ---------- ARTWORK PER CASE ----------
        elif header == "Artwork ($/SKU) =3800 ; Testing/QA=0 Per Case Calc ($)":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            check = isnumber_and("Annual Volume Cases QTY")
            cell.value = f"=IF({check},ROUND(3800/(2*{annual_vol}),2),0)" if annual_vol else "N/A"

        # ---------- ARTWORK % TO US FOB PICKUP ----------
        elif header == "Artwork % to US FOB Pickup price":
            artwork = ref_data_val("Artwork ($/SKU) =3800 ; Testing/QA=0 Per Case Calc ($)")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Artwork ($/SKU) =3800 ; Testing/QA=0 Per Case Calc ($)", "Total US FOB /Case pickup")
            cell.value = f"=IF({check},ROUND({artwork}/{total_us_fob_pickup}*100,2),0)" if (artwork and total_us_fob_pickup) else "N/A"

        # ---------- DOMESTIC COMPLIANCE PER CASE ----------
        elif header == "Domestic Compliance (qulaity assurance) = $500":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            check = isnumber_and("Annual Volume Cases QTY")
            cell.value = f"=IF({check},ROUND(500/{annual_vol},2),0)" if annual_vol else "N/A"

        # ---------- DOMESTIC COMPLIANCE % ----------
        elif header == "Domestic Compliance % to US FOB Pickup price":
            domestic = ref_data_val("Domestic Compliance (qulaity assurance) = $500")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Domestic Compliance (qulaity assurance) = $500", "Total US FOB /Case pickup")
            cell.value = f"=IF({check},ROUND(({domestic}/{total_us_fob_pickup})*100,2),0)" if (domestic and total_us_fob_pickup) else "N/A"

        # ---------- CERTIFICATION PER CASE ----------
        elif header == "Certification = $1000/facility (GFSI- food safety) per case for 2 Years":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            check = isnumber_and("Annual Volume Cases QTY")
            cell.value = f"=IF({check},ROUND(1000/(2*{annual_vol}),2),0)" if annual_vol else "N/A"

        # ---------- CERTIFICATION % ----------
        elif header == "Certification % to US FOB Pickup price":
            cert = ref_data_val("Certification = $1000/facility (GFSI- food safety) per case for 2 Years")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Certification = $1000/facility (GFSI- food safety) per case for 2 Years", "Total US FOB /Case pickup")
            cell.value = f"=IF({check},ROUND(({cert}/{total_us_fob_pickup})*100,2),0)" if (cert and total_us_fob_pickup) else "N/A"

        # ---------- PALLETIZING COST PER CASE ----------
        elif header.lower().startswith("palletizing cost per case"):
            cases_pallet = ref_data_val("Cases per pallet")
            check = isnumber_and("Cases per pallet")
            if cases_pallet:
                if Warehouse_inputs.upper() == "TAL":
                    cell.value = f"=IF({check},ROUND(50/{cases_pallet},2),0)"
                else:
                    cell.value = f"=IF({check},ROUND((14+12.5+5.75)/{cases_pallet},2),0)"
            else:
                cell.value = "N/A"

        # ---------- STORAGE COST PER CASE ----------
        elif header.lower().startswith("storage cost per case"):
            storage_months = ref_data_val("Storage Months")
            cases_per_pallet = ref_data_val("Cases per pallet")
            check = isnumber_and("Storage Months", "Cases per pallet")
            if storage_months and cases_per_pallet:
                if Warehouse_inputs.upper() == "TAL":
                    cell.value = f"=IF({check},ROUND((28*({storage_months}-1))/{cases_per_pallet},2),0)"
                else:
                    cell.value = f"=IF({check},ROUND((18.5*{storage_months})/{cases_per_pallet},2),0)"
            else:
                cell.value = "N/A"

        # ---------- WAREHOUSE % TO SUPPLIER FOB ----------
        elif header == "Warehouse % to supplier FOB price":
            pallet_cost = ref_one_of("Palletizing cost per case (TAL)", "Palletizing cost per case (ELM)")
            storage_cost = ref_one_of("Storage cost per case (TAL)", "Storage cost per case (ELM)")
            fob = ref_data_val("Finalized FOB/case")
            check = isnumber_and("Finalized FOB/case", PALLET_HDR, STORAGE_HDR)
            cell.value = f"=IF({check},ROUND((({pallet_cost}+{storage_cost})/{fob})*100,2),0)" if (pallet_cost and storage_cost and fob) else "N/A"

        # ---------- WAREHOUSE % TO US FOB PICKUP ----------
        elif header == "Warehouse % to US FOB pickup price":
            pallet_cost = ref_one_of("Palletizing cost per case (TAL)", "Palletizing cost per case (ELM)")
            storage_cost = ref_one_of("Storage cost per case (TAL)", "Storage cost per case (ELM)")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Total US FOB /Case pickup", PALLET_HDR, STORAGE_HDR)
            cell.value = f"=IF({check},ROUND((({pallet_cost}+{storage_cost})/{total_us_fob_pickup})*100,2),0)" if (pallet_cost and storage_cost and total_us_fob_pickup) else "N/A"

        # ---------- TOTAL/CASE ----------
        elif header == "Total/Case":
            chargeback = ref_one_of("Chargeback Fee", "Chargeback Fee = $")
            parts = [
                "Finalized FOB/case",
                "Duty/Case",
                "Inbound/case",
                "Artwork ($/SKU) =3800 ; Testing/QA=0 Per Case Calc ($)",
                "Domestic Compliance (qulaity assurance) = $500",
                "Certification = $1000/facility (GFSI- food safety) per case for 2 Years",
                PALLET_HDR,
                STORAGE_HDR,
            ]
            if chargeback:
                parts.append(chargeback)
            refs = [ref_data_val(p) if p not in (PALLET_HDR, STORAGE_HDR) else ref_one_of(PALLET_HDR, STORAGE_HDR) for p in parts if (ref_data_val(p) or p in (PALLET_HDR, STORAGE_HDR))]
            check = isnumber_and(*parts)
            cell.value = f"=IF({check},ROUND(" + "+".join(refs) + ",2),0)" if refs else "N/A"

        # ---------- FINANCING COST MONTHS ----------
        elif header == "Financing Cost Months (1 mth Prod + 1 mth ship + Storage Months)":
            storage_m = ref_data_val("Storage Months")
            check = isnumber_and("Storage Months")
            cell.value = f"=IF({check},ROUND(2+{storage_m},2),0)" if storage_m else "N/A"

        # ---------- OVERHEAD % ----------
        elif header == "Overhead Sales=1%, Daymon=0%; Supplier service fees- 0%; Payment Term (2%30 Net 30)= 3%; Delivery Issues= 1%; Reclamation= 0%, Financing Cost=0.5 % * Months":
            fin_m = ref_data_val("Financing Cost Months (1 mth Prod + 1 mth ship + Storage Months)")
            check = isnumber_and("Financing Cost Months (1 mth Prod + 1 mth ship + Storage Months)")
            cell.value = f"=IF({check},ROUND(0.05+0.005*{fin_m},2),0)" if fin_m else "N/A"

        # ---------- OVERHEAD PER CASE ----------
        elif header == "Overhead Per case":
            overhead_pct = ref_data_val("Overhead Sales=1%, Daymon=0%; Supplier service fees- 0%; Payment Term (2%30 Net 30)= 3%; Delivery Issues= 1%; Reclamation= 0%, Financing Cost=0.5 % * Months")
            total_case = ref_data_val("Total/Case")
            check = isnumber_and("Total/Case", "Overhead Sales=1%, Daymon=0%; Supplier service fees- 0%; Payment Term (2%30 Net 30)= 3%; Delivery Issues= 1%; Reclamation= 0%, Financing Cost=0.5 % * Months")
            cell.value = f"=IF({check},ROUND({overhead_pct}*{total_case},2),0)" if (overhead_pct and total_case) else "N/A"
        # ---------- OVERHEAD % TO US FOB PICKUP ----------
        elif header == "Overhead Per case % to US FOB pickup price":
            overhead_case = ref_data_val("Overhead Per case")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Overhead Per case", "Total US FOB /Case pickup")
            cell.value = f"=IF({check},ROUND(({overhead_case}/{total_us_fob_pickup})*100,2),0)" if (overhead_case and total_us_fob_pickup) else "N/A"

        # ---------- OVERHEAD TOTAL ($) ----------
        elif header == "Overhead Total ($)":
            overhead_case = ref_data_val("Overhead Per case")
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            check = isnumber_and("Overhead Per case", "Annual Volume Cases QTY")
            cell.value = f"=IF({check},ROUND({overhead_case}*{annual_vol},2),0)" if (overhead_case and annual_vol) else "N/A"

        # ---------- TOTAL US FOB/CASE ----------
        elif header == "Total US FOB/case":
            total_case = ref_data_val("Total/Case")
            overhead_case = ref_data_val("Overhead Per case")
            check = isnumber_and("Total/Case", "Overhead Per case")
            cell.value = f"=IF({check},ROUND({total_case}+{overhead_case},2),0)" if (total_case and overhead_case) else "N/A"

        # ---------- MARKUP/CASE ----------
        elif header == "Markup/Case":
            markup_pct = ref_data_val("Markup %")
            total_us_fob_case = ref_data_val("Total US FOB/case")
            check = isnumber_and("Markup %", "Total US FOB/case")
            cell.value = f"=IF({check},ROUND({markup_pct}/100*{total_us_fob_case},2),0)" if (markup_pct and total_us_fob_case) else "N/A"

        # ---------- TOTAL US FOB/CASE PICKUP ----------
        elif header == "Total US FOB /Case pickup":
            total_us_fob_case = ref_data_val("Total US FOB/case")
            markup_case = ref_data_val("Markup/Case")
            check = isnumber_and("Total US FOB/case", "Markup/Case")
            cell.value = f"=IF({check},ROUND({total_us_fob_case}+{markup_case},2),0)" if (total_us_fob_case and markup_case) else "N/A"

        # ---------- PACKS-PER-CASE-DEPENDENT FIELDS ----------
        elif header == "Total US FOB /Pack pickup":
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Total US FOB /Case pickup", "Finalized Supplier")

            ppc_expr = "(" + " + ".join([
                f'IF(ISNUMBER(SEARCH("{supplier_name.lower()}",LOWER(${finalized_supplier_col_letter}{formula_row}))),'
                f'IFERROR(VALUE({cols["packs_per_case"]}{formula_row}),0),0)'
                for supplier_name, cols in supplier_dimension_cols.items()
                if "packs_per_case" in cols
            ]) + ")"

            cell.value = f"=IF({check},ROUND(IFERROR({total_us_fob_pickup}/({ppc_expr}),0),2),0)" if total_us_fob_pickup else "N/A"

        elif header == "FTL Delivered pack price":
            base = ref_data_val("FTL Delivered case price")
            check = isnumber_and("FTL Delivered case price", "Finalized Supplier")

            ppc_expr = "(" + " + ".join([
                f'IF(ISNUMBER(SEARCH("{supplier_name.lower()}",LOWER(${finalized_supplier_col_letter}{formula_row}))),'
                f'IFERROR(VALUE({cols["packs_per_case"]}{formula_row}),0),0)'
                for supplier_name, cols in supplier_dimension_cols.items()
                if "packs_per_case" in cols
            ]) + ")"

            cell.value = f"=IF({check},ROUND(IFERROR({base}/({ppc_expr}),0),2),0)" if base else "N/A"

        elif header == "LTL Delivered pack price":
            base = ref_data_val("LTL Delivered case price")
            check = isnumber_and("LTL Delivered case price", "Finalized Supplier")

            ppc_expr = "(" + " + ".join([
                f'IF(ISNUMBER(SEARCH("{supplier_name.lower()}",LOWER(${finalized_supplier_col_letter}{formula_row}))),'
                f'IFERROR(VALUE({cols["packs_per_case"]}{formula_row}),0),0)'
                for supplier_name, cols in supplier_dimension_cols.items()
                if "packs_per_case" in cols
            ]) + ")"

            cell.value = f"=IF({check},ROUND(IFERROR({base}/({ppc_expr}),0),2),0)" if base else "N/A"

        elif header == "LTL Delivered PACK price for 3 pallets":
            base = ref_data_val("LTL Delivered case price for 3 pallets")
            check = isnumber_and("LTL Delivered case price for 3 pallets", "Finalized Supplier")

            ppc_expr = "(" + " + ".join([
                f'IF(ISNUMBER(SEARCH("{supplier_name.lower()}",LOWER(${finalized_supplier_col_letter}{formula_row}))),'
                f'IFERROR(VALUE({cols["packs_per_case"]}{formula_row}),0),0)'
                for supplier_name, cols in supplier_dimension_cols.items()
                if "packs_per_case" in cols
            ]) + ")"

            cell.value = f"=IF({check},ROUND(IFERROR({base}/({ppc_expr}),0),2),0)" if base else "N/A"

        # ---------- TCP NET PROFIT MARGIN ----------
        elif header == "TCP NET Profit Margin":
            markup_case = ref_data_val("Markup/Case")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Markup/Case", "Total US FOB /Case pickup")
            cell.value = f"=IF({check},ROUND(({markup_case}/{total_us_fob_pickup})*100,2),0)" if (markup_case and total_us_fob_pickup) else "N/A"

        # ---------- TOTAL % COMPONENTS ----------
        elif header == "TOTAL % components to US FOB pickup price":
            parts = [
                "Finalized FOB/ Case % to US FOB Pickup price",
                "Duty % to US FOB Pickup price",
                "Inbound % to US FOB Pickup price",
                "Artwork % to US FOB Pickup price",
                "Domestic Compliance % to US FOB Pickup price",
                "Certification % to US FOB Pickup price",
                "Warehouse % to US FOB pickup price",
                "Overhead Per case % to US FOB pickup price",
                "TCP NET Profit Margin",
            ]
            refs = [ref_data_val(p) for p in parts if ref_data_val(p)]
            check = isnumber_and(*parts)
            cell.value = f"=IF({check},ROUND(" + "+".join(refs) + ",2),0)" if refs else "N/A"
        # ---------- FTL / LTL CASE COUNTS ----------
        elif header == "No. of cases in 30 Pallets":
            cases_per_pallet = ref_data_val("Cases per pallet")
            check = isnumber_and("Cases per pallet")
            cell.value = f"=IF({check},ROUND(30*{cases_per_pallet},2),0)" if cases_per_pallet else "N/A"

        elif header == "No. of cases in 6 Pallets":
            cases_per_pallet = ref_data_val("Cases per pallet")
            check = isnumber_and("Cases per pallet")
            cell.value = f"=IF({check},ROUND(6*{cases_per_pallet},2),0)" if cases_per_pallet else "N/A"

        elif header == "No. of cases in 3 Pallets":
            cases_per_pallet = ref_data_val("Cases per pallet")
            check = isnumber_and("Cases per pallet")
            cell.value = f"=IF({check},ROUND(3*{cases_per_pallet},2),0)" if cases_per_pallet else "N/A"

        # ---------- FTL OUTFREIGHT / DELIVERED ----------
        elif header == "FTL outfreight/case":
            ftl_outfreight = ref_data_val("FTL Outfreight")
            cases_30 = ref_data_val("No. of cases in 30 Pallets")
            check = isnumber_and("FTL Outfreight","No. of cases in 30 Pallets")
            cell.value = f"=IF({check},ROUND({ftl_outfreight}/{cases_30},2),0)" if (ftl_outfreight and cases_30) else "N/A"

        elif header == "FTL Delivered case price":
            ftl_out_case = ref_data_val("FTL outfreight/case")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("FTL outfreight/case","Total US FOB /Case pickup")
            cell.value = f"=IF({check},ROUND({ftl_out_case}+{total_us_fob_pickup},2),0)" if (ftl_out_case and total_us_fob_pickup) else "N/A"

        elif header == "FTL Outfreight/FTL Delivered":
            ftl_out_case = ref_data_val("FTL outfreight/case")
            ftl_delivered_case = ref_data_val("FTL Delivered case price")
            check = isnumber_and("FTL outfreight/case","FTL Delivered case price")
            cell.value = f"=IF({check},ROUND(({ftl_out_case}/{ftl_delivered_case})*100,2),0)" if (ftl_out_case and ftl_delivered_case) else "N/A"

        # ---------- LTL (6 PALLETS) ----------
        elif header == "LTL Outfreight 6 pallets":
            ltl_outfreight = ref_data_val("LTL Outfreight")
            check = isnumber_and("LTL Outfreight")
            cell.value = f"=IF({check},ROUND({ltl_outfreight}*6,2),0)" if ltl_outfreight else "N/A"

        elif header == "LTL with Markup 15%":
            ltl_out_6 = ref_data_val("LTL Outfreight 6 pallets")
            check = isnumber_and("LTL Outfreight 6 pallets")
            cell.value = f"=IF({check},ROUND({ltl_out_6}*1.15,2),0)" if ltl_out_6 else "N/A"

        elif header == "LTL outfreight/case":
            ltl_with_markup = ref_data_val("LTL with Markup 15%")
            cases_6 = ref_data_val("No. of cases in 6 Pallets")
            check = isnumber_and("LTL with Markup 15%","No. of cases in 6 Pallets")
            cell.value = f"=IF({check},ROUND({ltl_with_markup}/{cases_6},2),0)" if (ltl_with_markup and cases_6) else "N/A"

        elif header == "LTL Delivered case price":
            ltl_out_case = ref_data_val("LTL outfreight/case")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("LTL outfreight/case","Total US FOB /Case pickup")
            cell.value = f"=IF({check},ROUND({ltl_out_case}+{total_us_fob_pickup},2),0)" if (ltl_out_case and total_us_fob_pickup) else "N/A"

        elif header == "LTL Outfreight/LTL Delivered":
            ltl_out_case = ref_data_val("LTL outfreight/case")
            ltl_delivered_case = ref_data_val("LTL Delivered case price")
            check = isnumber_and("LTL outfreight/case","LTL Delivered case price")
            cell.value = f"=IF({check},ROUND(({ltl_out_case}/{ltl_delivered_case})*100,2),0)" if (ltl_out_case and ltl_delivered_case) else "N/A"

        # ---------- LTL (3 PALLETS) ----------
        elif header == "LTL Outfreight 3 pallets":
            ltl_outfreight = ref_data_val("LTL Outfreight")
            check = isnumber_and("LTL Outfreight")
            cell.value = f"=IF({check},ROUND({ltl_outfreight}*3,2),0)" if ltl_outfreight else "N/A"

        elif header == "LTL with Markup 15% for 3 pallets":
            ltl_out_3 = ref_data_val("LTL Outfreight 3 pallets")
            check = isnumber_and("LTL Outfreight 3 pallets")
            cell.value = f"=IF({check},ROUND({ltl_out_3}*1.15,2),0)" if ltl_out_3 else "N/A"

        elif header == "LTL outfreight/case for 3 pallets":
            ltl_with_markup_3 = ref_data_val("LTL with Markup 15% for 3 pallets")
            cases_3 = ref_data_val("No. of cases in 3 Pallets")
            check = isnumber_and("LTL with Markup 15% for 3 pallets","No. of cases in 3 Pallets")
            cell.value = f"=IF({check},ROUND({ltl_with_markup_3}/{cases_3},2),0)" if (ltl_with_markup_3 and cases_3) else "N/A"

        elif header == "LTL Delivered case price for 3 pallets":
            ltl_out_case_3 = ref_data_val("LTL outfreight/case for 3 pallets")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("LTL outfreight/case for 3 pallets","Total US FOB /Case pickup")
            cell.value = f"=IF({check},ROUND({ltl_out_case_3}+{total_us_fob_pickup},2),0)" if (ltl_out_case_3 and total_us_fob_pickup) else "N/A"

        elif header == "LTL Outfreight/LTL Delivered for 3 pallets":
            ltl_out_case_3 = ref_data_val("LTL outfreight/case for 3 pallets")
            ltl_delivered_case_3 = ref_data_val("LTL Delivered case price for 3 pallets")
            check = isnumber_and("LTL outfreight/case for 3 pallets","LTL Delivered case price for 3 pallets")
            cell.value = f"=IF({check},ROUND(({ltl_out_case_3}/{ltl_delivered_case_3})*100,2),0)" if (ltl_out_case_3 and ltl_delivered_case_3) else "N/A"

        # ---------- REVENUE / PROFIT ----------
        elif header == "Total Revenue by Case (Pickup)":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            total_us_fob_pickup = ref_data_val("Total US FOB /Case pickup")
            check = isnumber_and("Annual Volume Cases QTY","Total US FOB /Case pickup")
            cell.value = f"=IF({check},ROUND({annual_vol}*{total_us_fob_pickup},2),0)" if (annual_vol and total_us_fob_pickup) else "N/A"

        elif header == "Net Total Profit (Pickup)":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            markup_case = ref_data_val("Markup/Case")
            check = isnumber_and("Annual Volume Cases QTY","Markup/Case")
            cell.value = f"=IF({check},ROUND({annual_vol}*{markup_case},2),0)" if (annual_vol and markup_case) else "N/A"

        elif header == "Total Revenue by Case (FTL)":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            ftl_delivered_case = ref_data_val("FTL Delivered case price")
            check = isnumber_and("Annual Volume Cases QTY","FTL Delivered case price")
            cell.value = f"=IF({check},ROUND({annual_vol}*{ftl_delivered_case},2),0)" if (annual_vol and ftl_delivered_case) else "N/A"

        elif header == "Net Total Profit (FTL)":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            ftl_delivered_case = ref_data_val("FTL Delivered case price")
            ftl_out_case = ref_data_val("FTL outfreight/case")
            total_us_fob_case = ref_data_val("Total US FOB/case")
            check = isnumber_and("Annual Volume Cases QTY","FTL Delivered case price","FTL outfreight/case","Total US FOB/case")
            cell.value = f"=IF({check},ROUND({annual_vol}*({ftl_delivered_case}-{ftl_out_case}-{total_us_fob_case}),2),0)" if (annual_vol and ftl_delivered_case and ftl_out_case and total_us_fob_case) else "N/A"

        elif header == "Total Revenue by Case (LTL) 6 pallets":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            ltl_delivered_case = ref_data_val("LTL Delivered case price")
            check = isnumber_and("Annual Volume Cases QTY","LTL Delivered case price")
            cell.value = f"=IF({check},ROUND({annual_vol}*{ltl_delivered_case},2),0)" if (annual_vol and ltl_delivered_case) else "N/A"

        elif header == "Net Total Profit (LTL) 6 pallets":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            ltl_delivered_case = ref_data_val("LTL Delivered case price")
            ltl_out_case = ref_data_val("LTL outfreight/case")
            total_us_fob_case = ref_data_val("Total US FOB/case")
            check = isnumber_and("Annual Volume Cases QTY","LTL Delivered case price","LTL outfreight/case","Total US FOB/case")
            cell.value = f"=IF({check},ROUND({annual_vol}*({ltl_delivered_case}-{ltl_out_case}-{total_us_fob_case}),2),0)" if (annual_vol and ltl_delivered_case and ltl_out_case and total_us_fob_case) else "N/A"

        elif header == "Total Revenue by Case (LTL) 3 pallets":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            ltl_delivered_case_3 = ref_data_val("LTL Delivered case price for 3 pallets")
            check = isnumber_and("Annual Volume Cases QTY","LTL Delivered case price for 3 pallets")
            cell.value = f"=IF({check},ROUND({annual_vol}*{ltl_delivered_case_3},2),0)" if (annual_vol and ltl_delivered_case_3) else "N/A"

        elif header == "Net Total Profit (LTL) 3 pallets":
            annual_vol = ref_data_val("Annual Volume Cases QTY")
            ltl_delivered_case_3 = ref_data_val("LTL Delivered case price for 3 pallets")
            ltl_out_case_3 = ref_data_val("LTL outfreight/case for 3 pallets")
            total_us_fob_case = ref_data_val("Total US FOB/case")
            check = isnumber_and("Annual Volume Cases QTY","LTL Delivered case price for 3 pallets","LTL outfreight/case for 3 pallets","Total US FOB/case")
            cell.value = f"=IF({check},ROUND({annual_vol}*({ltl_delivered_case_3}-{ltl_out_case_3}-{total_us_fob_case}),2),0)" if (annual_vol and ltl_delivered_case_3 and ltl_out_case_3 and total_us_fob_case) else "N/A"

        # ---------- CUSTOMER MARGIN CALCULATIONS ----------
        elif header == "Customer Margin % based on Pickup":
            selling_price = ref_data_val("Selling price")
            total_us_fob_pack = ref_data_val("Total US FOB /Pack pickup")
            check = isnumber_and("Selling price","Total US FOB /Pack pickup")
            cell.value = f"=IF({check},ROUND((({selling_price}-{total_us_fob_pack})/{selling_price})*100,2),0)" if (selling_price and total_us_fob_pack) else "N/A"

        elif header == "Customer Margin % based on FTL":
            selling_price = ref_data_val("Selling price")
            ftl_pack_price = ref_data_val("FTL Delivered pack price")
            check = isnumber_and("Selling price","FTL Delivered pack price")
            cell.value = f"=IF({check},ROUND((({selling_price}-{ftl_pack_price})/{selling_price})*100,2),0)" if (selling_price and ftl_pack_price) else "N/A"

        elif header == "Customer Margin % based on 6 pallet LTL":
            selling_price = ref_data_val("Selling price")
            ltl_pack_price = ref_data_val("LTL Delivered pack price")
            check = isnumber_and("Selling price","LTL Delivered pack price")
            cell.value = f"=IF({check},ROUND((({selling_price}-{ltl_pack_price})/{selling_price})*100,2),0)" if (selling_price and ltl_pack_price) else "N/A"

        elif header == "Customer Margin % based on 3 pallet LTL":
            selling_price = ref_data_val("Selling price")
            ltl_pack_price_3 = ref_data_val("LTL Delivered PACK price for 3 pallets")
            check = isnumber_and("Selling price","LTL Delivered PACK price for 3 pallets")
            cell.value = f"=IF({check},ROUND((({selling_price}-{ltl_pack_price_3})/{selling_price})*100,2),0)" if (selling_price and ltl_pack_price_3) else "N/A"

        else:
            cell.value = ""


        cell.fill = PatternFill(start_color="0cc8f7", end_color="0cc8f7", fill_type="solid")
        cell.font = Font(size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[formula_text_row].height = 30
    ws.row_dimensions[formula_row].height = 30

    current_row = formula_row + 1

    for row_idx in range(1, current_row):
        ws.row_dimensions[row_idx].height = 25

    # ------------------ Add Data Rows ------------------ #
    for ref_idx, ref in enumerate(all_refs):
        # Add TCP Number in first column
        tcp_num_cell = ws.cell(row=current_row, column=tcp_number_col, value=ref)
        tcp_num_cell.fill = PatternFill(start_color=data_row_color, end_color=data_row_color, fill_type="solid")
        tcp_num_cell.border = thin_border
        tcp_num_cell.alignment = Alignment(vertical="center")
       
        # Add TCP data if available (excluding reference column)
        if tcp_data:
            ref_key = str(ref).strip()
            tcp_row_data = tcp_data.get(ref_key, {})
            current_col = tcp_col_start

            for _, header in tcp_data_headers:
                # header is the column name string from the TCP sheet
                val = tcp_row_data.get(header, "")
                cell = ws.cell(row=current_row, column=current_col, value=val)
                cell.fill = PatternFill(start_color=data_row_color, end_color=data_row_color, fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                current_col += 1

        # Add supplier data columns
        for supplier_idx, (start_col, end_col, ref_col_idx, supplier_color, data_headers) in enumerate(column_positions):
            current_data_col = start_col
           
            headers, _, data = file_data[supplier_idx]
            if ref in data:
                row_data = data[ref]
                for header in data_headers:
                    header_idx = None
                    for i, orig_header in enumerate(headers):
                        if orig_header["value"] == header["value"]:
                            header_idx = i
                            break
                    if header_idx is not None and header_idx < len(row_data):
                        val = row_data[header_idx]["value"]
                    else:
                        val = ""
                    cell = ws.cell(row=current_row, column=current_data_col, value=val)
                    cell.fill = PatternFill(start_color=data_row_color, end_color=data_row_color, fill_type="solid")
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    current_data_col += 1
            else:
                for _ in data_headers:
                    cell = ws.cell(row=current_row, column=current_data_col, value="")
                    cell.fill = PatternFill(start_color=data_row_color, end_color=data_row_color, fill_type="solid")
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    current_data_col += 1

        # Add calculated results columns
        calc_data = calculation_results.get(ref, {})
        comp_data = comparison_results.get(ref, {})

        for idx, row_format in formatting_df.iterrows():
            header = row_format["Header"]
            value = calc_data.get(header, comp_data.get(header, ""))
            cell = ws.cell(row=current_row, column=calculated_col_start + idx, value=value)
            cell.fill = PatternFill(start_color=data_row_color, end_color=data_row_color, fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        current_row += 1

    # Adjust column widths
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
   
    # ------------------ DUPLICATE OUTPUT SHEET (STRICT SCHEMA FORMATTING) ------------------ #

    try:
        original_ws = ws
        copied_ws = wb.copy_worksheet(original_ws)
        copied_ws.title = "Merged Supplier Analysis - Copy"

        # --- Detect header row ---
        header_row = None
        for r in range(1, original_ws.max_row + 1):
            val = original_ws.cell(row=r, column=1).value
            if val and str(val).strip().lower() == "tcp number":
                header_row = r + 1
                break
        if not header_row:
            header_row = 4

        formula_row = header_row + 2
        data_start_row = formula_row
        max_row = copied_ws.max_row
        max_col = copied_ws.max_column

        # --- Explicit category lists (from your schema) ---
        percent_headers = [
            "Finalized FOB/ Case % to US FOB Pickup price",
            "Finalized Duty% based on finalized supplier",
            "General duty%",
            "Reciprocal %",
            "Section 301 China %",
            "Section 232 duty 25%",
            "Tariff %",
            "Total duty %",
            "Duty % to US FOB Pickup price",
            "Inbound % to supplier FOB price",
            "Inbound % to US FOB Pickup price",
            "Artwork % to US FOB Pickup price",
            "Domestic Compliance % to US FOB Pickup price",
            "Certification % to US FOB Pickup price",
            "Warehouse % to supplier FOB price",
            "Warehouse % to US FOB pickup price",
            "Overhead Sales=1%, Daymon=0%; Supplier service fees- 0%; Payment Term (2%30 Net 30)= 3%; Delivery Issues= 1%; Reclamation= 0%, Financing Cost=0.5 % * Months",
            "Overhead Per case % to US FOB pickup price",
            "Markup %",
            "TCP NET Profit Margin",
            "TOTAL % components to US FOB pickup price",
            "FTL Outfreight/FTL Delivered",
            "LTL Outfreight/LTL Delivered",
            "LTL Outfreight/LTL Delivered for 3 pallets",
            "Customer Margin % based on Pickup",
            "Customer Margin % based on FTL",
            "Customer Margin % based on 6 pallet LTL",
            "Customer Margin % based on 3 pallet LTL",
        ]

        dollar_headers = [
            "Finalized FOB/case",
            "General duty $",
            "Duty/Case",
            "Inbound to EAST Ocean freight",
            "Ocean per case",
            "Inbound to EAST (Ocean = $ 5500, Drayage = $1300, Unloading= $0/container)",
            "Inbound/case",
            "Artwork ($/SKU) =3800 ; Testing/QA=0 Per Case Calc ($)",
            "Domestic Compliance (qulaity assurance) = $500",
            "Certification = $1000/facility (GFSI- food safety) per case for 2 Years",
            "Palletizing cost per case (TAL)",
            "Palletizing cost per case (ELM)",
            "Storage cost per case (TAL)",
            "Storage cost per case (ELM)",
            "Chargeback Fee",
            "Total/Case",
            "Overhead Per case",
            "Overhead Total ($)",
            "Total US FOB/case",
            "Markup/Case",
            "Total US FOB /Case pickup",
            "Total US FOB /Pack pickup",
            "FTL Outfreight",
            "FTL outfreight/case",
            "FTL Delivered pack price",
            "FTL Delivered case price",
            "LTL Outfreight",
            "LTL Outfreight 6 pallets",
            "LTL with Markup 15%",
            "LTL outfreight/case",
            "LTL Delivered pack price",
            "LTL Delivered case price",
            "LTL Outfreight 3 pallets",
            "LTL with Markup 15% for 3 pallets",
            "LTL outfreight/case for 3 pallets",
            "LTL Delivered PACK price for 3 pallets",
            "LTL Delivered case price for 3 pallets",
            "Total Revenue by Case (Pickup)",
            "Net Total Profit (Pickup)",
            "Total Revenue by Case (FTL)",
            "Net Total Profit (FTL)",
            "Total Revenue by Case (LTL) 6 pallets",
            "Net Total Profit (LTL) 6 pallets",
            "Total Revenue by Case (LTL) 3 pallets",
            "Net Total Profit (LTL) 3 pallets",
            "Selling price"
        ]

        # --- Detect formula columns ---
        formula_columns = {}
        for col in range(1, max_col + 1):
            for row in range(1, formula_row + 1):
                val = original_ws.cell(row=row, column=col).value
                if isinstance(val, str) and val.strip().startswith("="):
                    header_val = original_ws.cell(row=header_row, column=col).value
                    formula_columns[col] = str(header_val).strip() if header_val else f"Column {col}"
                    break

        # --- Classify by explicit header match ---
        percent_cols, dollar_cols, numeric_cols = set(), set(), set()
        for col in range(1, max_col + 1):
            header_text = str(copied_ws.cell(row=header_row, column=col).value or "").strip()
            if header_text in dollar_headers:
                dollar_cols.add(col)
            elif header_text in percent_headers:
                percent_cols.add(col)
            else:
                numeric_cols.add(col)

        # --- Clear formula columns ---
        cleared = 0
        for col in formula_columns:
            for row in range(data_start_row, max_row + 1):
                c = copied_ws.cell(row=row, column=col)
                c.value = None
                c.data_type = "n"
                cleared += 1

        # --- Apply formats (no /100 conversion) ---
        for col in range(1, max_col + 1):
            if col in percent_cols:
                fmt = numbers.FORMAT_PERCENTAGE_00  # Display as percent
            elif col in dollar_cols:
                fmt = numbers.FORMAT_CURRENCY_USD_SIMPLE
            else:
                fmt = "0.00"

            for row in range(data_start_row, max_row + 200):
                c = copied_ws.cell(row=row, column=col)
                c.number_format = fmt
                c.alignment = Alignment(horizontal="right")

        # --- Normalize string constants ---
        def clean_value(cell):
            val = cell.value
            if isinstance(val, str):
                v = val.strip().replace(",", "")
                try:
                    if v.endswith("%"):
                        # ✅ Keep as numeric percent value (not divided by 100)
                        cell.value = float(v.replace("%", ""))/100
                        cell.number_format = numbers.FORMAT_PERCENTAGE_00
                    elif v.startswith("$"):
                        cell.value = float(v.replace("$", ""))
                        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                    else:
                        cell.value = float(v)
                except:
                    pass
            elif isinstance(val, (int, float)):
                if cell.column in percent_cols:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                elif cell.column in dollar_cols:
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                else:
                    cell.number_format = "0.00"

        retained = 0
        for row in copied_ws.iter_rows(min_row=data_start_row, max_row=max_row):
            for c in row:
                if c.value is not None:
                    clean_value(c)
                    retained += 1

        # --- Highlight cleared formula headers ---
        gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for col in formula_columns:
            copied_ws.cell(row=header_row, column=col).fill = gray

        # --- Log summary ---
        print("✅ Clean copy sheet created successfully.")
        print(f"   ➤ Cleared {len(formula_columns)} formula columns ({cleared:,} total cells).")
        print(f"   ➤ Retained {retained:,} constants (typed, formatted).")
        print(f"   ➤ % Columns = {len(percent_cols)}, $ Columns = {len(dollar_cols)}, Normal = {len(numeric_cols)}.")
        if formula_columns:
            print("   ➤ Cleared formula columns:\n      " + ", ".join(formula_columns.values()))

    except Exception as e:
        print(f"❌ Error creating clean copy sheet: {e}")

    return wb


def create_template():
    template_data = {
        'TCP Number': [],
        'Product Category': [],
        'Product Type': [],
        'TCP Description': [],
        'Pieces Per Case': [],
        'Packs Per Case': [],
        'Width inches': [],
        'Size (Sq.Ft/Inches)': [],
        'Material': [],
        'Color': [],
        'Thickness': [],
        'Features': [],
        'TI': [],
        'HI': [],
        'Annual Estimated Volume (Cases)': [],
        'Annual Estimated Volume (Container)': [],
        'Inner Packaging': [],
        'Outer Packaging': [],
        ' ': [],  
        'No Of Cases That Fit Into A Container': [],
        '  ': [],  
        'General Duty': [],
        'Reciprocal': [],
        'Section 301 China': [],
        'Section 232 duty': [],
        'Tariff': [],
        'Total Duty': [],
    }

    # Create DataFrame
    df = pd.DataFrame(template_data)
   
    # Create output stream
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Information from TrueChoicePack', startrow=1)  # shift down by 1 row

    workbook  = writer.book
    worksheet = writer.sheets['Information from TrueChoicePack']

    # ---------- Define header formats ----------
    header_format_yellow = workbook.add_format({'bold': True, 'bg_color': '#FFF2CC', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    header_format_green  = workbook.add_format({'bold': True, 'bg_color': '#C6E0B4', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    header_format_blue   = workbook.add_format({'bold': True, 'bg_color': '#BDD7EE', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    header_format_blank  = workbook.add_format({'bg_color': '#FFFFFF'})

    # 🔴 Red background + yellow text for the top duty section header
    duty_header_format = workbook.add_format({
        'bold': True, 'font_color': '#FFFF00', 'bg_color': '#C00000',
        'align': 'center', 'valign': 'vcenter', 'border': 1
    })

    # ---------- Identify index positions ----------
    first_group_end = list(template_data.keys()).index('Outer Packaging')
    second_group_col = list(template_data.keys()).index('No Of Cases That Fit Into A Container')
    second_space_col = list(template_data.keys()).index('  ')
    total_col_idx = list(template_data.keys()).index('Total Duty')

    # ---------- Write column headers ----------
    for idx, col in enumerate(df.columns):
        column_len = max(df[col].astype(str).map(len).max() if not df.empty else 0, len(col)) + 2
        worksheet.set_column(idx, idx, column_len)

        # Apply color formatting to each header
        if col.strip() == '':
            worksheet.write(1, idx, "", header_format_blank)
        elif idx <= first_group_end:
            worksheet.write(1, idx, col, header_format_yellow)
        elif idx == second_group_col:
            worksheet.write(1, idx, col, header_format_green)
        elif idx > second_space_col:
            worksheet.write(1, idx, col, header_format_blue)
        else:
            worksheet.write(1, idx, col, header_format_blank)

    # ---------- Formula for Total Duty ----------
    for row in range(3, 1002):  # adjusted because we started headers at row 2
        start_col = chr(65 + total_col_idx - 5)
        end_col   = chr(65 + total_col_idx - 1)
        formula = f'=IF(SUM({start_col}{row}:{end_col}{row})=0,"",SUM({start_col}{row}:{end_col}{row}))'
        worksheet.write_formula(row-1, total_col_idx, formula)

    # ---------- Add merged heading ONLY above the duty section ----------
    duty_start_col = list(template_data.keys()).index('General Duty')
    duty_end_col = list(template_data.keys()).index('Total Duty')

    worksheet.merge_range(
        0, duty_start_col, 0, duty_end_col,
        "Only fill the duty for the selected country",
        duty_header_format
    )

    # ---------- Adjust row height for better visibility ----------
    worksheet.set_row(0, 25)  # merged header row
    worksheet.set_row(1, 20)  # main column header row

    writer.close()
    output.seek(0)
    return output



def create_supplier_template():
    # Column Names
    columns = [
        "Sr.No.", "TCP Number", "Item Category", "Product Description", "Pieces Per Pack",
        "Packs per Case", "Total Pcs per case", "Size (OZ)", "Item Color", "Item weight (grams)",
        "Item Dimension (mm)", "Double Wall or Single Wall", "Material",
        "BPI Logo Printing on the product (individual product)", "Qualities/Claims",
        "Recyclable Packaging Required", "Retail Pack Length (mm)", "Retail Pack Width (mm)",
        "Retail Pack Height (mm)", "Retail Pack Weight (grams)",
        "Annual Volume Packs QTY", "Annual Volume Cases QTY",
        "Primary Packaging", "Secondary / Outer Packaging",
        "End of Info from TCP", "Carton Length (inches)", "Carton Width (inches)",
        "Carton Height (inches)", "Carton Weight (lbs.)", "Per pcs Dimension (inch)",
        "Per pcs Weight (grams)", "# cartons can fit into a 40 HQ Container?",
        "FOB Port name", "FOB Price per Pack", "FOB Price per Case",
        "HTS Code", "Duty % to USA",
        "Do you have existing tooling of this item? Yes/No",
        "If not, please confirm the tooling cost per set.",
        "If you don't have existing tooling, to meet our demand. How many sets of tooling are required?",
        "Available Product Certifications (Kindly attach the certificates also with the mail)",
        "Do you have BPI Certification? (Yes/No) If yes, do provide documentation",
        "Do you have TUV Industrial Compost Certification? (Yes/No) If yes, do provide documentation",
        "Do you have Ok HOME Compost Certification? (Yes/No) If yes, do provide documentation",
        "Do you have USDA Certification? (Yes/No) If yes, do provide documentation",
        "Comments (if any additional information need to provide)"
    ]

    df = pd.DataFrame(columns=columns)
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Supplier Information')

    workbook = writer.book
    worksheet = writer.sheets['Supplier Information']

    # 🎨 Color Styles
    colors = {
        "blue": "#D9E1F2",
        "peach": "#FCE4D6",
        "pink_red": "#F8CBCB",
        "red": "#C00000",
        "green": "#E2EFDA",
        "yellow": "#FFF200",
        "skyblue": "#CFE2F3",
    }

    # Create formatting styles
    fmt_blue = workbook.add_format({'bold': True, 'bg_color': colors['blue'], 'border': 1,
                                    'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_peach = workbook.add_format({'bold': True, 'bg_color': colors['peach'], 'border': 1,
                                     'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_pink_red = workbook.add_format({'bold': True, 'bg_color': colors['pink_red'], 'font_color': '#FF0000',
                                        'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_red = workbook.add_format({'bold': True, 'bg_color': colors['red'], 'font_color': '#FFFFFF',
                                   'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_green = workbook.add_format({'bold': True, 'bg_color': colors['green'], 'border': 1,
                                     'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_yellow = workbook.add_format({'bold': True, 'bg_color': colors['yellow'], 'border': 1,
                                      'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_skyblue = workbook.add_format({'bold': True, 'bg_color': colors['skyblue'], 'border': 1,
                                       'align': 'center', 'valign': 'vcenter', 'text_wrap': True})

    # --- Dynamic Color Rules ---
    def pick_color(col):
        c = col.lower()
        if "annual volume" in c:
            return fmt_pink_red
        elif "end of info" in c:
            return fmt_red
        elif "recyclable" in c:
            return fmt_peach
        elif "fob price" in c:
            return fmt_yellow
        elif any(k in c for k in ["fob", "carton", "duty", "hts", "port"]):
            return fmt_green
        elif any(k in c for k in ["certification", "bpi", "tuv", "ok home", "usda", "comment"]):
            return fmt_skyblue
        else:
            return fmt_blue

    # --- Apply color formats dynamically ---
    for idx, col in enumerate(columns):
        fmt = pick_color(col)
        worksheet.write(0, idx, col, fmt)
        worksheet.set_column(idx, idx, max(18, len(col) * 0.9))

    worksheet.freeze_panes(1, 0)
    writer.close()
    output.seek(0)
    return output



def _copy_sheet_with_styles(src_ws, dst_ws):
    """
    Safely copy all cell values, formulas, styles, merged cells, column widths,
    and row heights from src_ws → dst_ws, skipping MergedCell placeholders.
    """
    # 1️⃣ Copy cell values + styles
    for row in src_ws.iter_rows():
        for cell in row:
            # Skip merged-cell placeholders (they throw AttributeError)
            if cell.__class__.__name__ == "MergedCell":
                continue

            tgt = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            if cell.has_style:
                tgt.font = _copy(cell.font)
                tgt.border = _copy(cell.border)
                tgt.fill = _copy(cell.fill)
                tgt.number_format = cell.number_format
                tgt.protection = _copy(cell.protection)
                tgt.alignment = _copy(cell.alignment)

    # 2️⃣ Copy merged ranges
    if src_ws.merged_cells.ranges:
        for mcr in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(mcr))

    # 3️⃣ Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dst_ws.column_dimensions[col_letter].width = dim.width

    # 4️⃣ Copy row heights
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst_ws.row_dimensions[row_idx].height = dim.height

    # 5️⃣ Copy sheet view (optional)
    try:
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    except Exception:
        pass


def download_workbook(wb):
    """
    Merge the generated workbook into a macro-enabled template and let the user download a .xlsm
    that includes VBA + button + your data (both 'Merged Supplier Analysis' and its '- Copy').
    """
    macro_template = os.path.join(os.getcwd(), "TCP_Macro_Template.xlsm")
    if not os.path.exists(macro_template):
        st.error("⚠️ Macro template not found. Place 'TCP_Macro_Template.xlsm' in the app folder.")
        return

    # 1) Save your generated workbook to a temp .xlsx
    temp_xlsx = os.path.join(os.getcwd(), "Merged_Supplier_Analysis_Temp.xlsx")
    wb.save(temp_xlsx)

    # 2) Load the generated workbook (with formulas) and the macro template (keep_vba=True)
    data_wb = load_workbook(temp_xlsx, data_only=False, keep_vba=False)
    macro_wb = load_workbook(macro_template, keep_vba=True)

    # 3) Remove all sheets from the macro template (openpyxl needs at least one during the process)
    #    We'll keep one placeholder and delete it after we copy.
    placeholder = macro_wb.active
    placeholder_title = placeholder.title

    # 4) Copy sheets over
    for src_ws in data_wb.worksheets:
        # create the destination sheet with the same name
        dst_ws = macro_wb.create_sheet(title=src_ws.title)
        _copy_sheet_with_styles(src_ws, dst_ws)

    # 5) Delete the placeholder (template’s original sheet)
    #    (only if it still exists and is not one of the desired sheets)
    if placeholder_title in [ws.title for ws in macro_wb.worksheets]:
        # ensure there's at least one other sheet before deleting
        if len(macro_wb.worksheets) > 1:
            del macro_wb[placeholder_title]

    # 6) Save macro-enabled result
    out_path = os.path.join(os.getcwd(), "Merged_Supplier_Analysis_Output.xlsm")
    macro_wb.save(out_path)

    # 7) Offer download
    st.success("✅ Macro-enabled workbook ready!")
    with open(out_path, "rb") as f:
        st.download_button(
            label="⬇️ Download Full Analysis (Macro Enabled)",
            data=f,
            file_name="Merged_Supplier_Analysis_Output.xlsm",
            mime="application/vnd.ms-excel.sheet.macroEnabled.12",
        )


def Pricing_analysis_V2():
    st.title("Supplier Pricing Analysis")

    st.markdown("Download RFQ template for supplier")
    supplier_template = create_supplier_template()
    st.download_button(
            label="Download RFQ Template",
            data=supplier_template,
            file_name="RFQ_Sheet_For_Supplier.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.markdown("Download TCP Information template for analytics")
    template = create_template()
    st.download_button(
            label="Download TCP Information Template",
            data=template,
            file_name="Information_from_TrueChoicePack.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    TCP_info_uploaded_file = TCP_info_upload_files()
    if not TCP_info_uploaded_file:
        return
   
    uploaded_files = upload_files()
    if not uploaded_files:
        return

    try:
        tcp_inputs = get_tcp_inputs(context="min_supplier")
        if st.button("Run Full Analysis"):
            try:
                # Merge files including TCP data
                all_refs, file_data, title_rows, file_names, tcp_headers, tcp_ref_idx, tcp_data = merge_file_by_reference(
                    uploaded_files,
                    TCP_info_uploaded_file
                )

                # Run comparison to find min FOB supplier
                comparison_results = comparison_function(all_refs, file_data, file_names)
                min_fob_results = {ref: comp_data["Min Supplier"] for ref, comp_data in comparison_results.items()}
               
                # Calculate pricing for min FOB supplier
                calculation_results = calculation_for_min_fob_supplier(
                    uploaded_files,
                    TCP_info_uploaded_file,
                    tcp_inputs,
                    min_fob_results
                )

                # Create workbook with TCP info merged by reference
                wb_min_fob = create_workbook(
                    all_refs,
                    file_data,
                    title_rows,
                    file_names,
                    calculation_results,
                    tcp_inputs,
                    comparison_results,
                    tcp_headers=tcp_headers,
                    tcp_ref_idx=tcp_ref_idx,
                    tcp_data=tcp_data
                )

                st.success("Analysis completed successfully!")
                download_workbook(wb_min_fob)

            except Exception as e:
                st.error(f"Processing error: {str(e)}")
                import traceback
                st.error(traceback.format_exc())

    except Exception as e:
        st.error(f"Initialization error: {str(e)}")
        import traceback
        st.error(traceback.format_exc())

# ----------------- MAIN APP ROUTING ----------------- #
if selected_option == "Pricing Analysis Version - 2":
    Pricing_analysis_V2()
elif selected_option == "Home":
    st.title("Pricing Analysis Dashboard")
    st.markdown("""
        Welcome to the Supplier Pricing Analysis Tool.  

        - Compare pricing across multiple suppliers  
        - Merge TCP information into supplier reports **by reference number**
        - Calculate Total Cost of Procurement (TCP)  
        - Generate comprehensive reports  
       
        ### Key Features:
        - **TCP Info Integration**: TCP data is now matched with supplier data using the same reference number
        - **Reference-Based Merging**: All data (TCP + Suppliers) aligned by TCP/Reference Number
        - **Automated Calculations**: Pricing formulas automatically applied
        - **Excel Export**: Download complete analysis with all sections
    """)